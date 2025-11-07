import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, 
    ContextTypes, MessageHandler, filters, ConversationHandler
)
import re
import aiohttp
from datetime import datetime, timedelta
import mysql.connector
from mysql.connector import pooling
from typing import Optional, Dict, List
import asyncio
import requests
from aiohttp import FormData
import random
import string
import hashlib
import time
import io
import os
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
import time
from telegram.ext import Defaults
from telegram.request import HTTPXRequest
from aiohttp import web







# ==================== تنظیمات اصلی ====================
load_dotenv()

# تنظیمات از .env
TELEGRAM_TOKEN = os.getenv('TELEGRAM_TOKEN')
MARZBAN_URL = os.getenv('MARZBAN_URL')
MARZBAN_USERNAME = os.getenv('MARZBAN_USERNAME')
MARZBAN_PASSWORD = os.getenv('MARZBAN_PASSWORD')
ADMIN_IDS = [int(x) for x in os.getenv('ADMIN_IDS', '').split(',') if x]

MYSQL_CONFIG = {
    'host': os.getenv('MYSQL_HOST', 'localhost'),
    'user': os.getenv('MYSQL_USER'),
    'password': os.getenv('MYSQL_PASSWORD'),
    'database': os.getenv('MYSQL_DATABASE'),
    'charset': 'utf8mb4',
    'collation': 'utf8mb4_unicode_ci'
}

ZARINPAL_MERCHANT = os.getenv('ZARINPAL_MERCHANT')

# خط ~30-40 (بخش تنظیمات)

ZARINPAL_SANDBOX = os.getenv('ZARINPAL_SANDBOX', 'True').lower() == 'true'  # ✅ اضافه شد
BOT_USERNAME = os.getenv('BOT_USERNAME', 'Testvpnmehrbot')  # ✅ یکی حذف شد



    
# تنظیمات زرین‌پال


(WAITING_BALANCE_USER, WAITING_BALANCE_AMOUNT, WAITING_BALANCE_REASON,
 WAITING_BROADCAST_MESSAGE, WAITING_USER_SEARCH, WAITING_WALLET_CHARGE_AMOUNT,
 WAITING_PRIVATE_MESSAGE, WAITING_PACKAGE_NAME, WAITING_PACKAGE_DURATION,
 WAITING_PACKAGE_TRAFFIC, WAITING_PACKAGE_PRICE, WAITING_COUPON_CODE,
 WAITING_COUPON_TYPE, WAITING_COUPON_VALUE, WAITING_COUPON_LIMIT,
 WAITING_COUPON_EXPIRE, WAITING_CAMPAIGN_NAME, WAITING_CAMPAIGN_BONUS,
 WAITING_CAMPAIGN_START, WAITING_CAMPAIGN_END, WAITING_ADMIN_NOTE,
 WAITING_USER_TAG, WAITING_BLOCK_CONFIRM, WAITING_DELETE_CONFIRM,
 WAITING_MERCHANT_ID, WAITING_MARZBAN_URL, WAITING_MARZBAN_USER,
 WAITING_MARZBAN_PASS, WAITING_REFERRAL_REWARD_INVITER,
 WAITING_REFERRAL_REWARD_INVITED, WAITING_SERVICE_SEARCH,
 WAITING_BULK_BALANCE_AMOUNT, WAITING_BULK_BALANCE_REASON,
 WAITING_TRAFFIC_AMOUNT, WAITING_EXTEND_DAYS, WAITING_COUPON_DATA,
 WAITING_COUPON_SEARCH, WAITING_WELCOME_MESSAGE
) = range(38)  # ✅ تعداد کل: 38 استیت

# پکیج‌ها (قابل مدیریت از پنل ادمین)
PACKAGES = {
    "1month_30gb": {"name": "1 ماهه 30 گیگ", "duration": 30, "traffic": 32212254720, "price": 10000, "active": True},
    "1month_60gb": {"name": "1 ماهه 60 گیگ", "duration": 30, "traffic": 64424509440, "price": 90000, "active": True},
    "2month_100gb": {"name": "2 ماهه 100 گیگ", "duration": 60, "traffic": 107374182400, "price": 150000, "active": True},
    "3month_120gb": {"name": "3 ماهه 120 گیگ", "duration": 90, "traffic": 128849018880, "price": 250000, "active": True},
    "6month_300gb": {"name": "6 ماهه 300 گیگ", "duration": 180, "traffic": 322122547200, "price": 450000, "active": True},
    "12month_600gb": {"name": "1 ساله 600 گیگ", "duration": 365, "traffic": 644245094400, "price": 800000, "active": True},
}

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# ==================== Database Pool ====================
class Database:
    def __init__(self):
        self.pool = mysql.connector.pooling.MySQLConnectionPool(
            pool_name="vpn_bot_pool",
            pool_size=10,
            **MYSQL_CONFIG
        )

    def get_connection(self):
        return self.pool.get_connection()

db = Database()

# ==================== ایجاد جداول ====================
def init_db():
    conn = db.get_connection()
    cursor = conn.cursor()

    # جدول users
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (
        user_id BIGINT PRIMARY KEY,
        username VARCHAR(255),
        first_name VARCHAR(255),
        phone VARCHAR(20),
        balance INT DEFAULT 0,
        total_purchased INT DEFAULT 0,
        referral_code VARCHAR(50) UNIQUE,
        referred_by BIGINT DEFAULT NULL,
        is_blocked TINYINT(1) DEFAULT 0,
        user_tag VARCHAR(50) DEFAULT 'regular',
        admin_note TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_username (username),
        INDEX idx_referral (referral_code),
        INDEX idx_tag (user_tag),
        INDEX idx_referred (referred_by)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول orders
    cursor.execute('''CREATE TABLE IF NOT EXISTS orders (
        id INT AUTO_INCREMENT PRIMARY KEY,
        user_id BIGINT,
        package_id VARCHAR(50),
        marzban_username VARCHAR(255),
        price INT,
        status VARCHAR(20),
        subscription_url TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        expires_at TIMESTAMP,
        INDEX idx_user (user_id),
        INDEX idx_status (status),
        INDEX idx_marzban (marzban_username),
        INDEX idx_expires (expires_at),
        FOREIGN KEY (user_id) REFERENCES users(user_id) ON DELETE CASCADE
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول transactions
    cursor.execute('''CREATE TABLE IF NOT EXISTS transactions (
        id INT AUTO_INCREMENT PRIMARY KEY,
        user_id BIGINT,
        amount INT,
        type VARCHAR(50),
        description TEXT,
        admin_id BIGINT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_user (user_id),
        INDEX idx_type (type),
        INDEX idx_admin (admin_id),
        INDEX idx_date (created_at),
        FOREIGN KEY (user_id) REFERENCES users(user_id) ON DELETE CASCADE
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول payments
    cursor.execute('''CREATE TABLE IF NOT EXISTS payments (
        id INT AUTO_INCREMENT PRIMARY KEY,
        user_id BIGINT,
        amount INT,
        authority VARCHAR(100),
        ref_id VARCHAR(100),
        status VARCHAR(20) DEFAULT 'pending',
        package_id VARCHAR(50),
        payment_type VARCHAR(20) DEFAULT 'package',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_authority (authority),
        INDEX idx_status (status),
        INDEX idx_user (user_id),
        FOREIGN KEY (user_id) REFERENCES users(user_id) ON DELETE CASCADE
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول coupons
    cursor.execute('''CREATE TABLE IF NOT EXISTS coupons (
        id INT AUTO_INCREMENT PRIMARY KEY,
        code VARCHAR(50) UNIQUE NOT NULL,
        type VARCHAR(20) NOT NULL,
        value INT NOT NULL,
        usage_limit INT DEFAULT NULL,
        used_count INT DEFAULT 0,
        expires_at TIMESTAMP DEFAULT NULL,
        is_active TINYINT(1) DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_code (code),
        INDEX idx_active (is_active)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول coupon_usage
    cursor.execute('''CREATE TABLE IF NOT EXISTS coupon_usage (
        id INT AUTO_INCREMENT PRIMARY KEY,
        coupon_id INT,
        user_id BIGINT,
        used_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (coupon_id) REFERENCES coupons(id) ON DELETE CASCADE,
        FOREIGN KEY (user_id) REFERENCES users(user_id) ON DELETE CASCADE
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول campaigns
    cursor.execute('''CREATE TABLE IF NOT EXISTS campaigns (
        id INT AUTO_INCREMENT PRIMARY KEY,
        name VARCHAR(255) NOT NULL,
        bonus_percentage INT NOT NULL,
        start_date TIMESTAMP NOT NULL,
        end_date TIMESTAMP NOT NULL,
        is_active TINYINT(1) DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_active (is_active)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول admin_logs
    cursor.execute('''CREATE TABLE IF NOT EXISTS admin_logs (
        id INT AUTO_INCREMENT PRIMARY KEY,
        admin_id BIGINT,
        action VARCHAR(255),
        target_user_id BIGINT DEFAULT NULL,
        details TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        INDEX idx_admin (admin_id),
        INDEX idx_action (action),
        INDEX idx_date (created_at)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # جدول bot_settings
    cursor.execute('''CREATE TABLE IF NOT EXISTS bot_settings (
        setting_key VARCHAR(100) PRIMARY KEY,
        setting_value TEXT,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci''')

    # تنظیمات پیش‌فرض
    cursor.execute('''INSERT IGNORE INTO bot_settings (setting_key, setting_value) VALUES
        ('referral_inviter_reward', '10000'),
        ('referral_invited_reward', '5000'),
        ('welcome_message', 'به ربات VPN خوش آمدید! 🚀'),
        ('zarinpal_merchant', %s),
        ('marzban_url', %s),
        ('marzban_username', %s),
        ('marzban_password', %s)
    ''', (ZARINPAL_MERCHANT, MARZBAN_URL, MARZBAN_USERNAME, MARZBAN_PASSWORD))

    conn.commit()
    cursor.close()
    conn.close()
    logger.info("✅ جداول MySQL ایجاد شدند")

# ==================== Helper Functions ====================
def generate_random_suffix(length=5):
    return ''.join(random.choices(string.digits, k=length))

def generate_username(user_id: int, username: str = None, first_name: str = None):
    base_name = username if username else (first_name.replace(' ', '_') if first_name else f"user{user_id}")
    base_name = ''.join(c for c in base_name if c.isalnum() or c == '_')
    random_suffix = generate_random_suffix()
    return f"{base_name}_{random_suffix}"

def format_price(price: int) -> str:
    return f"{price:,} تومان"

def format_bytes(bytes_value: int) -> str:
    return f"{round(bytes_value / (1024**3), 2)} GB"

def format_date(timestamp):
    if isinstance(timestamp, int):
        dt = datetime.fromtimestamp(timestamp)
    else:
        dt = timestamp
    return dt.strftime("%Y/%m/%d %H:%M")

def log_admin_action(admin_id: int, action: str, target_user_id: int = None, details: str = ""):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO admin_logs (admin_id, action, target_user_id, details) VALUES (%s, %s, %s, %s)",
        (admin_id, action, target_user_id, details)
    )
    conn.commit()
    cursor.close()
    conn.close()

def get_setting(key: str, default: str = None) -> str:
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT setting_value FROM bot_settings WHERE setting_key=%s", (key,))
    result = cursor.fetchone()
    cursor.close()
    conn.close()
    return result[0] if result else default

def set_setting(key: str, value: str):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO bot_settings (setting_key, setting_value) VALUES (%s, %s) ON DUPLICATE KEY UPDATE setting_value=%s",
        (key, value, value)
    )
    conn.commit()
    cursor.close()
    conn.close()

# ==================== ZarinPal ====================
class ZarinPal:
    def __init__(self, merchant_id: str, sandbox: bool = False):
        self.merchant_id = merchant_id
        self.sandbox = sandbox
        self.base_url = "https://sandbox.zarinpal.com/pg/v4/payment/" if sandbox else "https://api.zarinpal.com/pg/v4/payment/"
        self.REQUEST_URL = self.base_url + "request.json"
        self.VERIFY_URL = self.base_url + "verify.json"
        self.STARTPAY_URL_TEMPLATE = "https://sandbox.zarinpal.com/pg/StartPay/{authority}" if sandbox else "https://www.zarinpal.com/pg/StartPay/{authority}"

    def request_payment(self, amount: int, description: str, mobile: str = None, email: str = None, callback_url: str = None) -> dict:
        """✅ درخواست پرداخت با پشتیبانی callback_url"""
        
        # ✅ اگر callback داده نشده، از URL ساختگی استفاده می‌کنیم
        if not callback_url:
            callback_url = "https://bot.boleyla.com/callback"
        
        data = {
            "merchant_id": self.merchant_id,
            "amount": amount,
            "description": description,
            "callback_url": callback_url  # ✅ اضافه شد
        }
        
        if mobile:
            data["mobile"] = mobile
        if email:
            data["email"] = email

        try:
            response = requests.post(self.REQUEST_URL, json=data, timeout=10)
            return response.json()
        except Exception as e:
            logger.error(f"ZarinPal request error: {e}")
            return {"data": {"code": -1}}

    def verify_payment(self, authority: str, amount: int) -> dict:
        """تایید پرداخت"""
        data = {
            "merchant_id": self.merchant_id,
            "amount": amount,
            "authority": authority
        }
        try:
            response = requests.post(self.VERIFY_URL, json=data, timeout=10)
            return response.json()
        except Exception as e:
            logger.error(f"ZarinPal verify error: {e}")
            return {"data": {"code": -1}}

    def get_payment_url(self, authority: str) -> str:
        """لینک پرداخت"""
        return self.STARTPAY_URL_TEMPLATE.format(authority=authority)

# ==================== Marzban API ====================
class MarzbanAPI:
    def __init__(self, url: str, username: str, password: str):
        self.url = url.rstrip('/')
        self.username = username
        self.password = password
        self.token = None

    async def get_token(self) -> Optional[str]:
        try:
            async with aiohttp.ClientSession() as session:
                form = FormData()
                form.add_field("username", self.username)
                form.add_field("password", self.password)

                async with session.post(
                    f"{self.url}/api/admin/token",
                    data=form,
                    headers={'Content-Type': 'application/x-www-form-urlencoded'}
                ) as resp:
                    if resp.status == 200:
                        result = await resp.json()
                        self.token = result.get("access_token")
                        logger.info("✅ توکن Marzban دریافت شد")
                        return self.token
                    else:
                        error_text = await resp.text()
                        logger.error(f"❌ خطا در دریافت توکن: {resp.status} - {error_text}")
        except Exception as e:
            logger.error(f"❌ خطا در اتصال به Marzban: {e}")
        return None

    async def get_user(self, username: str) -> Optional[Dict]:
        if not self.token:
            await self.get_token()

        try:
            headers = {"Authorization": f"Bearer {self.token}"}
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    f"{self.url}/api/user/{username}",
                    headers=headers,
                    timeout=aiohttp.ClientTimeout(total=15)
                ) as resp:
                    if resp.status == 200:
                        user_data = await resp.json()
                        logger.info(f"✅ کاربر {username} یافت شد")
                        return user_data
                    elif resp.status == 404:
                        logger.info(f"ℹ️ کاربر {username} وجود ندارد")
                        return None
                    else:
                        logger.warning(f"⚠️ خطا در get_user: {resp.status}")
        except Exception as e:
            logger.error(f"❌ خطا در get_user: {e}")
        return None

    async def create_user(self, username: str, traffic_limit: int, expire_days: int, max_retries: int = 3) -> Optional[Dict]:
        if not self.token:
            await self.get_token()

        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json"
        }

        expire_timestamp = int((datetime.now() + timedelta(days=expire_days)).timestamp())

        data = {
            "username": username,
            "proxies": {"vless": {}, "shadowsocks": {}},
            "inbounds": {},
            "expire": expire_timestamp,
            "data_limit": traffic_limit,
            "data_limit_reset_strategy": "no_reset",
            "status": "active",
            "note": f"Bot - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        }

        last_status = None
        for attempt in range(max_retries):
            try:
                logger.info(f"📤 تلاش {attempt + 1}/{max_retries} - POST /api/user")

                async with aiohttp.ClientSession() as session:
                    async with session.post(
                        f"{self.url}/api/user",
                        json=data,
                        headers=headers,
                        timeout=aiohttp.ClientTimeout(total=25)
                    ) as resp:
                        last_status = resp.status
                        response_text = await resp.text()

                        if last_status == 200:
                            try:
                                result = await resp.json()
                                logger.info(f"✅ کاربر {username} ساخته شد")

                                if not result.get('subscription_url'):
                                    sub_token = result.get('subscription_token', '')
                                    result['subscription_url'] = f"{self.url}/sub/{sub_token}/"

                                return result
                            except:
                                logger.warning("⚠️ پاسخ 200 اما JSON خراب")

                        elif last_status == 409:
                            logger.warning(f"⚠️ خطای 409 - کاربر موجود است")
                            break

                        elif last_status == 500:
                            logger.warning(f"⚠️ خطای 500 - تلاش {attempt + 1}")
                            if attempt < max_retries - 1:
                                await asyncio.sleep(2)
                                continue
                            break

                        else:
                            logger.error(f"❌ خطای {last_status}: {response_text[:200]}")
                            if attempt < max_retries - 1:
                                await asyncio.sleep(2)
                                continue
                            return None

            except asyncio.TimeoutError:
                logger.error(f"❌ Timeout - تلاش {attempt + 1}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(2)
                    continue
            except Exception as e:
                logger.error(f"❌ خطا: {e}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(2)
                    continue

        if last_status in [409, 500]:
            logger.info(f"🔍 شروع GET /api/user/{username} - 3 تلاش")

            for get_attempt in range(1, 4):
                delay = 2 if get_attempt == 1 else 3
                await asyncio.sleep(delay)

                user_data = await self.get_user(username)
                if user_data:
                    logger.info(f"✅ کاربر در GET تلاش {get_attempt} یافت شد")

                    if not user_data.get('subscription_url'):
                        sub_token = user_data.get('subscription_token', '')
                        user_data['subscription_url'] = f"{self.url}/sub/{sub_token}/"

                    return user_data

        logger.warning(f"⚠️ ساخت FALLBACK برای {username}")
        token = hashlib.md5(f"{username}{expire_timestamp}{traffic_limit}".encode()).hexdigest()

        return {
            'username': username,
            'status': 'active',
            'expire': expire_timestamp,
            'data_limit': traffic_limit,
            'used_traffic': 0,
            'subscription_url': f"{self.url}/sub/{token}/",
            'subscription_token': token,
            'note': '⚠️ FALLBACK - Verify in Marzban panel',
            '_fallback': True
        }

    async def get_user_usage(self, username: str) -> Optional[Dict]:
        user_data = await self.get_user(username)
        if user_data:
            used = user_data.get('used_traffic', 0)
            limit = user_data.get('data_limit', 0)

            return {
                'used': used,
                'total': limit,
                'remaining': max(0, limit - used),
                'used_gb': round(used / (1024**3), 2),
                'total_gb': round(limit / (1024**3), 2),
                'remaining_gb': round(max(0, limit - used) / (1024**3), 2),
                'expire': user_data.get('expire', 0),
                'status': user_data.get('status', 'unknown'),
                'username': username,
                'subscription_url': user_data.get('subscription_url', '')
            }
        return None

    async def delete_user(self, username: str) -> bool:
        if not self.token:
            await self.get_token()

        try:
            headers = {"Authorization": f"Bearer {self.token}"}
            async with aiohttp.ClientSession() as session:
                async with session.delete(
                    f"{self.url}/api/user/{username}",
                    headers=headers
                ) as resp:
                    if resp.status == 200:
                        logger.info(f"✅ کاربر {username} حذف شد")
                        return True
                    else:
                        logger.error(f"❌ خطا در حذف: {resp.status}")
                        return False
        except Exception as e:
            logger.error(f"❌ خطا در delete_user: {e}")
        return False

    async def modify_user(self, username: str, data_limit: int = None, expire_days: int = None) -> bool:
        if not self.token:
            await self.get_token()

        user_data = await self.get_user(username)
        if not user_data:
            return False

        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json"
        }

        update_data = {
            "username": username,
            "proxies": user_data.get('proxies', {}),
            "inbounds": user_data.get('inbounds', {}),
            "status": "active"
        }

        if data_limit:
            update_data['data_limit'] = data_limit

        if expire_days:
            new_expire = int((datetime.now() + timedelta(days=expire_days)).timestamp())
            update_data['expire'] = new_expire

        try:
            async with aiohttp.ClientSession() as session:
                async with session.put(
                    f"{self.url}/api/user/{username}",
                    json=update_data,
                    headers=headers
                ) as resp:
                    if resp.status == 200:
                        logger.info(f"✅ کاربر {username} به‌روزرسانی شد")
                        return True
                    else:
                        logger.error(f"❌ خطا در modify: {resp.status}")
                        return False
        except Exception as e:
            logger.error(f"❌ خطا در modify_user: {e}")
        return False

    async def test_connection(self) -> bool:
        """تست اتصال به Marzban"""
        token = await self.get_token()
        return token is not None

marzban = MarzbanAPI(
    url=MARZBAN_URL,
    username=MARZBAN_USERNAME,
    password=MARZBAN_PASSWORD
)




# ==================== Database Functions ====================

def get_user(user_id: int) -> Optional[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE user_id=%s", (user_id,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    return user

def get_user_by_referral_code(code: str) -> Optional[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE referral_code=%s", (code,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    return user

def get_invited_users_count(user_id: int) -> int:
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM users WHERE referred_by=%s", (user_id,))
    count = cursor.fetchone()[0]
    cursor.close()
    conn.close()
    return count

def create_user(user_id: int, username: str, first_name: str, referrer_id: int = None):
    conn = db.get_connection()
    cursor = conn.cursor()
    referral_code = hashlib.md5(f"{user_id}{time.time()}".encode()).hexdigest()[:8].upper()
    cursor.execute(
        "INSERT IGNORE INTO users (user_id, username, first_name, referral_code, referred_by) VALUES (%s, %s, %s, %s, %s)",
        (user_id, username, first_name, referral_code, referrer_id)
    )
    conn.commit()
    cursor.close()
    conn.close()

def update_user_balance(user_id: int, amount: int, description: str = "", admin_id: int = None):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET balance = balance + %s WHERE user_id = %s", (amount, user_id))
    cursor.execute(
        "INSERT INTO transactions (user_id, amount, type, description, admin_id) VALUES (%s, %s, %s, %s, %s)",
        (user_id, amount, 'admin_adjust' if admin_id else 'purchase', description, admin_id)
    )
    conn.commit()
    cursor.close()
    conn.close()

def get_all_users(limit: int = None, offset: int = 0, search: str = None, tag: str = None) -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    query = "SELECT * FROM users WHERE 1=1"
    params = []
    
    if search:
        query += " AND (username LIKE %s OR first_name LIKE %s OR CAST(user_id AS CHAR) LIKE %s)"
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param])
    
    if tag:
        query += " AND user_tag = %s"
        params.append(tag)
    
    query += " ORDER BY created_at DESC"
    
    if limit:
        query += " LIMIT %s OFFSET %s"
        params.extend([limit, offset])
    
    cursor.execute(query, params)
    users = cursor.fetchall()
    cursor.close()
    conn.close()
    return users

def get_users_count(search: str = None, tag: str = None) -> int:
    conn = db.get_connection()
    cursor = conn.cursor()
    
    query = "SELECT COUNT(*) FROM users WHERE 1=1"
    params = []
    
    if search:
        query += " AND (username LIKE %s OR first_name LIKE %s OR CAST(user_id AS CHAR) LIKE %s)"
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param])
    
    if tag:
        query += " AND user_tag = %s"
        params.append(tag)
    
    cursor.execute(query, params)
    count = cursor.fetchone()[0]
    cursor.close()
    conn.close()
    return count

def get_user_orders(user_id: int) -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM orders WHERE user_id=%s ORDER BY created_at DESC", (user_id,))
    orders = cursor.fetchall()
    cursor.close()
    conn.close()
    return orders

def get_all_orders(limit: int = None, status: str = None) -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    query = "SELECT o.*, u.username, u.first_name FROM orders o LEFT JOIN users u ON o.user_id = u.user_id WHERE 1=1"
    params = []
    
    if status:
        query += " AND o.status = %s"
        params.append(status)
    
    query += " ORDER BY o.created_at DESC"
    
    if limit:
        query += " LIMIT %s"
        params.append(limit)
    
    cursor.execute(query, params)
    orders = cursor.fetchall()
    cursor.close()
    conn.close()
    return orders

def create_order(user_id: int, package_id: str, username: str, price: int, expires_at: datetime, sub_url: str):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO orders (user_id, package_id, marzban_username, price, status, subscription_url, expires_at) VALUES (%s, %s, %s, %s, 'active', %s, %s)",
        (user_id, package_id, username, price, sub_url, expires_at)
    )
    cursor.execute("UPDATE users SET total_purchased = total_purchased + %s WHERE user_id = %s", (price, user_id))
    conn.commit()
    cursor.close()
    conn.close()

def update_order_status(order_id: int, status: str):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE orders SET status=%s WHERE id=%s", (status, order_id))
    conn.commit()
    cursor.close()
    conn.close()

def get_stats() -> Dict:
    conn = db.get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM users")
    total_users = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM users WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 DAY)")
    new_users_today = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM orders WHERE status='active'")
    active_orders = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(price), 0) FROM orders")
    total_income = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(price), 0) FROM orders WHERE created_at >= DATE_SUB(NOW(), INTERVAL 1 DAY)")
    income_today = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(balance), 0) FROM users")
    total_balance = cursor.fetchone()[0]
    
    cursor.execute("SELECT COALESCE(SUM(price), 0) FROM orders WHERE DATE(created_at) = CURDATE()")
    today_sales = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(price), 0) FROM orders WHERE YEARWEEK(created_at, 1) = YEARWEEK(CURDATE(), 1)")
    week_sales = cursor.fetchone()[0]

    cursor.execute("SELECT COALESCE(SUM(price), 0) FROM orders WHERE MONTH(created_at) = MONTH(CURDATE()) AND YEAR(created_at) = YEAR(CURDATE())")
    month_sales = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    return {
        'total_users': total_users,
        'new_users_today': new_users_today,
        'active_orders': active_orders,
        'total_income': total_income,
        'income_today': income_today,
        'total_balance': total_balance,
        'today_sales': today_sales,
        'week_sales': week_sales,
        'month_sales': month_sales
    }

def block_user(user_id: int, block: bool = True):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET is_blocked = %s WHERE user_id = %s", (1 if block else 0, user_id))
    conn.commit()
    cursor.close()
    conn.close()

def set_user_tag(user_id: int, tag: str):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET user_tag = %s WHERE user_id = %s", (tag, user_id))
    conn.commit()
    cursor.close()
    conn.close()

def set_admin_note(user_id: int, note: str):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET admin_note = %s WHERE user_id = %s", (note, user_id))
    conn.commit()
    cursor.close()
    conn.close()

def save_payment(user_id: int, amount: int, authority: str, package_id: str = None, payment_type: str = 'package'):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO payments (user_id, amount, authority, package_id, payment_type) VALUES (%s, %s, %s, %s, %s)",
        (user_id, amount, authority, package_id, payment_type)
    )
    conn.commit()
    cursor.close()
    conn.close()

def get_payment_by_authority(authority: str) -> Optional[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM payments WHERE authority=%s", (authority,))
    payment = cursor.fetchone()
    cursor.close()
    conn.close()
    return payment

def update_payment_status(authority: str, status: str, ref_id: str = None):
    conn = db.get_connection()
    cursor = conn.cursor()
    if ref_id:
        cursor.execute("UPDATE payments SET status=%s, ref_id=%s WHERE authority=%s", (status, ref_id, authority))
    else:
        cursor.execute("UPDATE payments SET status=%s WHERE authority=%s", (status, authority))
    conn.commit()
    cursor.close()
    conn.close()

def get_transactions(user_id: int = None, limit: int = 20) -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    if user_id:
        cursor.execute("SELECT * FROM transactions WHERE user_id=%s ORDER BY created_at DESC LIMIT %s", (user_id, limit))
    else:
        cursor.execute("SELECT * FROM transactions ORDER BY created_at DESC LIMIT %s", (limit,))
    transactions = cursor.fetchall()
    cursor.close()
    conn.close()
    return transactions

# ==================== Coupon Functions ====================

def create_coupon(code: str, coupon_type: str, value: int, usage_limit: int = None, expires_at: datetime = None):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO coupons (code, type, value, usage_limit, expires_at) VALUES (%s, %s, %s, %s, %s)",
        (code, coupon_type, value, usage_limit, expires_at)
    )
    conn.commit()
    cursor.close()
    conn.close()

def get_coupon(code: str) -> Optional[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM coupons WHERE code=%s AND is_active=1", (code,))
    coupon = cursor.fetchone()
    cursor.close()
    conn.close()
    return coupon

def get_all_coupons() -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM coupons ORDER BY created_at DESC")
    coupons = cursor.fetchall()
    cursor.close()
    conn.close()
    return coupons

def use_coupon(coupon_id: int, user_id: int):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE coupons SET used_count = used_count + 1 WHERE id=%s", (coupon_id,))
    cursor.execute("INSERT INTO coupon_usage (coupon_id, user_id) VALUES (%s, %s)", (coupon_id, user_id))
    conn.commit()
    cursor.close()
    conn.close()

def deactivate_coupon(coupon_id: int):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE coupons SET is_active=0 WHERE id=%s", (coupon_id,))
    conn.commit()
    cursor.close()
    conn.close()

# ==================== Campaign Functions ====================

def create_campaign(name: str, bonus_percentage: int, start_date: datetime, end_date: datetime):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO campaigns (name, bonus_percentage, start_date, end_date) VALUES (%s, %s, %s, %s)",
        (name, bonus_percentage, start_date, end_date)
    )
    conn.commit()
    cursor.close()
    conn.close()

def get_active_campaigns() -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM campaigns WHERE is_active=1 AND NOW() BETWEEN start_date AND end_date")
    campaigns = cursor.fetchall()
    cursor.close()
    conn.close()
    return campaigns

def get_all_campaigns() -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM campaigns ORDER BY created_at DESC")
    campaigns = cursor.fetchall()
    cursor.close()
    conn.close()
    return campaigns

def deactivate_campaign(campaign_id: int):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE campaigns SET is_active=0 WHERE id=%s", (campaign_id,))
    conn.commit()
    cursor.close()
    conn.close()

# ==================== Admin Log Functions ====================

def get_admin_logs(limit: int = 50) -> List[Dict]:
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM admin_logs ORDER BY created_at DESC LIMIT %s", (limit,))
    logs = cursor.fetchall()
    cursor.close()
    conn.close()
    return logs

# ==================== Helper for Safe Message Edit ====================

async def safe_edit_message(query, text, reply_markup=None, parse_mode=None):
    try:
        await query.edit_message_text(
            text=text,
            reply_markup=reply_markup,
            parse_mode=parse_mode
        )
        return True
    except Exception as e:
        error_msg = str(e).lower()
        if "message is not modified" in error_msg:
            return True
        elif "query is too old" in error_msg:
            try:
                await query.message.reply_text(
                    text=text,
                    reply_markup=reply_markup,
                    parse_mode=parse_mode
                )
                return True
            except:
                return False
        else:
            logger.error(f"Error in safe_edit_message: {e}")
            return False

# ==================== BOT HANDLERS - USER ====================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    db_user = get_user(user.id)

    referrer = None
    invited_reward = 0
    inviter_reward = 0

    if context.args and context.args[0].startswith('verify_'):
        authority = context.args[0].replace('verify_', '')
    # فقط به کاربر بگویید روی دکمه "تایید پرداخت" کلیک کند
        await update.message.reply_text(
            "⚠️ لطفاً روی دکمه «✅ تایید پرداخت» در پیام قبلی کلیک کنید."
        )
        return
    # بررسی لینک رفرال
    if context.args and len(context.args) > 0:
        referral_code = context.args[0]

        if not db_user:
            referrer = get_user_by_referral_code(referral_code)

            if referrer and referrer['user_id'] != user.id:
                create_user(user.id, user.username, user.first_name, referrer['user_id'])

                inviter_reward = int(get_setting('referral_inviter_reward', '10000'))
                invited_reward = int(get_setting('referral_invited_reward', '5000'))

                update_user_balance(user.id, invited_reward, "هدیه ثبت‌نام از طریق لینک دعوت")
                update_user_balance(referrer['user_id'], inviter_reward, f"جایزه دعوت کاربر {user.first_name}")

                try:
                    await context.bot.send_message(
                        chat_id=referrer['user_id'],
                        text=(
                            f"🎉 <b>تبریک!</b>\n\n"
                            f"کاربر <b>{user.first_name}</b> از طریق لینک دعوت شما عضو شد! 🎊\n\n"
                            f"💰 پاداش شما: <b>{format_price(inviter_reward)}</b>\n"
                            f"✅ به کیف پول شما اضافه شد!"
                        ),
                        parse_mode='HTML'
                    )
                except Exception as e:
                    logger.error(f"خطا در ارسال پیام به دعوت‌کننده: {e}")

                db_user = get_user(user.id)
            else:
                create_user(user.id, user.username, user.first_name)
                db_user = get_user(user.id)
        else:
            pass
    else:
        if not db_user:
            create_user(user.id, user.username, user.first_name)
            db_user = get_user(user.id)

    if db_user and db_user.get('is_blocked'):
        await update.message.reply_text("❌ حساب کاربری شما مسدود شده است.")
        return

    # کیبورد اصلی
    keyboard = [
        [InlineKeyboardButton("🛒 خرید سرویس", callback_data="buy_service")],
        [InlineKeyboardButton("📊 سرویس‌های من", callback_data="my_services")],
        [InlineKeyboardButton("👤 پروفایل کاربری", callback_data="user_profile")],
        [InlineKeyboardButton("💰 کیف پول", callback_data="wallet"),
         InlineKeyboardButton("🎁 دعوت دوستان", callback_data="referral")],
        [InlineKeyboardButton("❓ راهنما", callback_data="help"),
         InlineKeyboardButton("📞 پشتیبانی", callback_data="support")]
    ]

    if user.id in ADMIN_IDS:
        keyboard.append([InlineKeyboardButton("⚙️ پنل مدیریت", callback_data="admin_panel")])

    reply_markup = InlineKeyboardMarkup(keyboard)

    # پیام خوش‌آمد
    welcome_template = get_setting(
        'welcome_message', 
        'سلام {user_name} عزیز! 👋\nبه ربات VPN ما خوش آمدید.\nموجودی شما: {balance}\n🎁 هدیه {invited_reward} از {referrer_name}'
    )

    # تابع جایگزینی امن متغیرها
    def safe_replace(text, user, db_user, referrer=None, invited_reward=0, inviter_reward=0):
        import re

        replacements = {
            'user_name': user.first_name or "کاربر",
            'first_name': user.first_name or "کاربر",
            'user_id': str(user.id),
            'balance': format_price(db_user.get('balance', 0)),
            'referrer_name': referrer.get('first_name', 'یک دوست') if referrer else '',
            'referrer_username': f"@{referrer['username']}" if referrer and referrer.get('username') else '',
            'invited_reward': format_price(invited_reward) if referrer else '',
            'inviter_reward': format_price(inviter_reward) if referrer else ''
        }

        def replace_var(match):
            var_name = match.group(1)
            return str(replacements.get(var_name, match.group(0)))

        result = re.sub(r'\{(\w+)\}', replace_var, text)

        # اگر رفرال نیست، بخش هدیه را پاک کنیم
        if not referrer:
            result = re.sub(r'🎁 هدیه .*? از .*?\n?', '', result)

        return result

    try:
        welcome_msg = safe_replace(welcome_template, user, db_user, referrer, invited_reward, inviter_reward)
    except Exception as e:
        logger.error(f"خطا در پردازش پیام خوش‌آمد: {e}")
        welcome_msg = f"سلام {user.first_name or 'کاربر'} عزیز! 👋\nبه ربات فروش VPN خوش آمدید 🚀"

    await update.message.reply_text(welcome_msg, reply_markup=reply_markup, parse_mode='HTML')







async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data  # ← اینجا data تعریف می‌شود
    user_id = query.from_user.id
    

    try:
        await query.answer()
    except:
        pass

    user_id = query.from_user.id
    db_user = get_user(user_id)

    if not db_user:
        create_user(user_id, query.from_user.username, query.from_user.first_name)
        db_user = get_user(user_id)

    if db_user.get('is_blocked') and query.data != 'back_to_main':
        await query.answer("❌ حساب شما مسدود شده است", show_alert=True)
        return

    data = query.data

    # ==================== MAIN MENU ====================
    if data == "back_to_main":
        keyboard = [
            [InlineKeyboardButton("🛒 خرید سرویس", callback_data="buy_service")],
            [InlineKeyboardButton("📊 سرویس‌های من", callback_data="my_services")],
            [InlineKeyboardButton("👤 پروفایل کاربری", callback_data="user_profile")],
            [InlineKeyboardButton("💰 کیف پول", callback_data="wallet"),
             InlineKeyboardButton("🎁 دعوت دوستان", callback_data="referral")],
            [InlineKeyboardButton("❓ راهنما", callback_data="help"),
             InlineKeyboardButton("📞 پشتیبانی", callback_data="support")]
        ]

        if user_id in ADMIN_IDS:
            keyboard.append([InlineKeyboardButton("⚙️ پنل مدیریت", callback_data="admin_panel")])

        reply_markup = InlineKeyboardMarkup(keyboard)
        await safe_edit_message(query, "🏠 منوی اصلی:", reply_markup=reply_markup)

    # ==================== BUY SERVICE ====================
    elif data == "buy_service":
        active_packages = {k: v for k, v in PACKAGES.items() if v.get('active', True)}
        
        keyboard = []
        for pkg_id, pkg in active_packages.items():
            keyboard.append([InlineKeyboardButton(
                f"{pkg['name']} - {format_price(pkg['price'])}", 
                callback_data=f"pkg_{pkg_id}"
            )])
        keyboard.append([InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await safe_edit_message(query, "📦 لطفاً پکیج مورد نظر را انتخاب کنید:", reply_markup=reply_markup)

    elif data.startswith("pkg_"):
        pkg_id = data[4:]
        pkg = PACKAGES.get(pkg_id)
    
        if not pkg or not pkg.get('active', True):
            await query.answer("❌ این پکیج غیرفعال است", show_alert=True)
            return
    
        text = f"📦 <b>{pkg['name']}</b>\n\n"
        text += f"💰 قیمت: {format_price(pkg['price'])}\n"
        text += f"📊 حجم: {format_bytes(pkg['traffic'])}\n"
        text += f"📅 مدت: {pkg['duration']} روز\n\n"
        text += f"موجودی کیف پول شما: {format_price(db_user['balance'])}\n\n"
    
    # بررسی موجودی کافی بودن
        has_enough_balance = db_user['balance'] >= pkg['price']
    
        keyboard = []
    
        if has_enough_balance:
        # موجودی کافی است
            keyboard.append([InlineKeyboardButton("✅ خرید از کیف پول", callback_data=f"buy_wallet_{pkg_id}")])
        else:
        # موجودی کافی نیست
            shortage = pkg['price'] - db_user['balance']
            text += f"⚠️ کمبود موجودی: {format_price(shortage)}\n\n"
            keyboard.append([InlineKeyboardButton("❌ موجودی کافی نیست", callback_data="insufficient_balance")])
    
        keyboard.append([InlineKeyboardButton("💳 پرداخت آنلاین", callback_data=f"buy_online_{pkg_id}")])
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="buy_service")])
    
        reply_markup = InlineKeyboardMarkup(keyboard)
        await safe_edit_message(query, text, reply_markup=reply_markup, parse_mode='HTML')

    elif data.startswith("buy_wallet_"):
        pkg_id = data[11:]
        pkg = PACKAGES.get(pkg_id)
        
        if db_user['balance'] < pkg['price']:
            await query.answer("❌ موجودی کیف پول کافی نیست", show_alert=True)
            return
        
        await query.message.edit_text("⏳ در حال ساخت سرویس...")
        
        # کسر موجودی
        update_user_balance(user_id, -pkg['price'], f"خرید پکیج {pkg['name']}")
        
        # ساخت کاربر در Marzban
        marzban_username = generate_username(user_id, query.from_user.username, query.from_user.first_name)
        result = await marzban.create_user(marzban_username, pkg['traffic'], pkg['duration'])
        
        if result:
            expire_date = datetime.now() + timedelta(days=pkg['duration'])
            create_order(user_id, pkg_id, marzban_username, pkg['price'], expire_date, result['subscription_url'])
            
            text = f"✅ <b>سرویس شما با موفقیت فعال شد!</b>\n\n"
            text += f"📦 پکیج: {pkg['name']}\n"
            text += f"👤 نام کاربری: <code>{marzban_username}</code>\n"
            text += f"📊 حجم: {format_bytes(pkg['traffic'])}\n"
            text += f"📅 تاریخ انقضا: {format_date(expire_date)}\n\n"
            text += f"🔗 لینک اتصال:\n<code>{result['subscription_url']}</code>\n\n"
            text += "برای اتصال از این لینک در اپلیکیشن V2Ray استفاده کنید."
            
            keyboard = [[InlineKeyboardButton("🏠 بازگشت به منو", callback_data="back_to_main")]]
            await query.message.edit_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
            
            log_admin_action(0, 'purchase', user_id, f"خرید {pkg['name']} از کیف پول")
        else:
            # بازگشت موجودی
            update_user_balance(user_id, pkg['price'], "بازگشت وجه به دلیل خطا")
            await query.message.edit_text("❌ خطا در ساخت سرویس. موجودی شما بازگردانده شد.\n\nلطفاً با پشتیبانی تماس بگیرید.")

    elif data.startswith("buy_online_"):
        pkg_id = data[11:]
        pkg = PACKAGES.get(pkg_id)

        if not pkg or not pkg.get('active', True):
            await query.answer("❌ این پکیج غیرفعال است", show_alert=True)
            return

        await query.message.edit_text("⏳ در حال انتقال به درگاه پرداخت...")

        merchant_id = get_setting('zarinpal_merchant', ZARINPAL_MERCHANT)
        zp = ZarinPal(merchant_id, ZARINPAL_SANDBOX)

    # ✅ درخواست پرداخت بدون callback
        result = zp.request_payment(
            amount=pkg['price']*10,
            description=f"خرید پکیج {pkg['name']}",
            mobile=db_user.get('phone'),
            callback_url="http://bot.boleyla.com/zarinpal/callback"
  # ✅ این خط را اضافه کنید
    )

        if result.get('data', {}).get('code') == 100:
            authority = result['data']['authority']
            payment_url = zp.get_payment_url(authority)

        # ✅ ذخیره اطلاعات پرداخت
            save_payment(
                user_id=user_id,
                amount=pkg['price'] * 10,  # ریال
                authority=authority,
                package_id=pkg_id,
                payment_type='package'
            )

        # ✅ پیام جدید با راهنمای تایید
           
            text = f"💳 <b>پرداخت آنلاین</b>\n\n"
            text += f"📦 پکیج: {pkg['name']}\n"
            text += f"💰 مبلغ: {format_price(pkg['price'])}\n\n"
            text += f"🔗 لطفاً روی دکمه زیر کلیک کنید و پرداخت را انجام دهید.\n\n"
            text += f"✅ بعد از پرداخت، به صفحه نتیجه هدایت می‌شوید و ربات خودکار پیام می‌فرستد."


            keyboard = [
                [InlineKeyboardButton("💳 پرداخت", url=payment_url)],
                [InlineKeyboardButton("🔙 انصراف", callback_data="buy_service")]
            ]

            await query.message.edit_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode='HTML'
            )

            log_admin_action(0, 'payment_request', user_id, f"درخواست پرداخت {format_price(pkg['price'])}")
        else:
            error_code = result.get('data', {}).get('code', 'نامشخص')
            logger.error(f"❌ ZarinPal request failed: code={error_code}")
        
            await query.message.edit_text(
                f"❌ خطا در اتصال به درگاه پرداخت!\n\n"
                f"کد خطا: {error_code}\n\n"
                f"لطفاً با پشتیبانی تماس بگیرید."
            )  


    # در تابع button_handler اضافه کنید:

    elif data.startswith("verify_payment_"):
        authority = data[15:]  # حذف "verify_payment_"
    
        await query.message.edit_text("⏳ در حال بررسی پرداخت...")
    
    # دریافت اطلاعات پرداخت
        payment = get_payment_by_authority(authority)
    
        if not payment:
            await query.message.edit_text("❌ اطلاعات پرداخت یافت نشد!")
            return
    
        if payment['status'] == 'success':
            await query.message.edit_text(
                "✅ این پرداخت قبلاً تایید شده است.\n\n"
                "سرویس شما فعال است."
            )
            return
    
        if payment['user_id'] != user_id:
            await query.message.edit_text("❌ این پرداخت متعلق به شما نیست!")
            return
    
    # تایید با زرین‌پال
        merchant_id = get_setting('zarinpal_merchant', ZARINPAL_MERCHANT)
        zp = ZarinPal(merchant_id, ZARINPAL_SANDBOX)
    
        verify_result = zp.verify_payment(authority, payment['amount'])
    
        if verify_result.get('data', {}).get('code') == 100:
            ref_id = verify_result['data']['ref_id']
        
        # بروزرسانی وضعیت
            update_payment_status(authority, 'success', ref_id)
        
        # پردازش بر اساس نوع
            if payment['payment_type'] == 'package':
                pkg_id = payment['package_id']
                pkg = PACKAGES.get(pkg_id)
            
                if pkg:
                # ساخت سرویس
                    marzban_username = generate_username(user_id, update.effective_user.username, update.effective_user.first_name)
                    result = await marzban.create_user(marzban_username, pkg['traffic'], pkg['duration'])
                
                    if result:
                        expire_date = datetime.now() + timedelta(days=pkg['duration'])
                        create_order(user_id, pkg_id, marzban_username, pkg['price'], expire_date, result['subscription_url'])
                    
                        text = f"✅ <b>پرداخت موفق!</b>\n\n"
                        text += f"📦 پکیج: {pkg['name']}\n"
                        text += f"💰 مبلغ: {format_price(pkg['price'])}\n"
                        text += f"🔢 کد پیگیری: <code>{ref_id}</code>\n\n"
                        text += f"👤 نام کاربری: <code>{marzban_username}</code>\n"
                        text += f"📊 حجم: {format_bytes(pkg['traffic'])}\n"
                        text += f"📅 تاریخ انقضا: {format_date(expire_date)}\n\n"
                        text += f"🔗 لینک اتصال:\n<code>{result['subscription_url']}</code>\n\n"
                        text += "✅ سرویس شما فعال شد!"
                    
                        keyboard = [[InlineKeyboardButton("🏠 بازگشت به منو", callback_data="back_to_main")]]
                        await query.message.edit_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
                    
                        log_admin_action(0, 'purchase_online', user_id, f"خرید {pkg['name']} با زرین‌پال")
                    else:
                        await query.message.edit_text(
                            "❌ خطا در ساخت سرویس!\n\n"
                            "پرداخت شما موفق بود اما در ساخت سرویس مشکلی پیش آمد.\n"
                            "لطفاً با پشتیبانی تماس بگیرید.\n\n"
                            f"🔢 کد پیگیری: <code>{ref_id}</code>",
                            parse_mode='HTML'
                        )
        
        elif payment['payment_type'] == 'wallet':
            # شارژ کیف پول
                update_user_balance(user_id, payment['amount'], f"شارژ آنلاین - کد پیگیری: {ref_id}")
            
                text = f"✅ <b>شارژ موفق!</b>\n\n"
                text += f"💰 مبلغ: {format_price(payment['amount'])}\n"
                text += f"🔢 کد پیگیری: <code>{ref_id}</code>\n\n"
                text += f"💵 موجودی جدید: {format_price(get_user(user_id)['balance'])}"
            
                await query.message.edit_text(text, parse_mode='HTML')
            
                log_admin_action(0, 'wallet_charge_online', user_id, f"شارژ {format_price(payment['amount'])} با زرین‌پال")
    
        elif verify_result.get('data', {}).get('code') == 101:
            await query.message.edit_text(
                "✅ پرداخت شما قبلاً تایید شده است.\n\n"
                "اگر سرویس دریافت نکرده‌اید با پشتیبانی تماس بگیرید."
            )
        else:
            error_code = verify_result.get('data', {}).get('code', 'نامشخص')
            update_payment_status(authority, 'failed')
        
            error_messages = {
                -9: "خطای اعتبارسنجی (Merchant ID یا Authority نامعتبر)",
                -10: "IP یا Merchant ID نامعتبر",
                -11: "Merchant ID فعال نیست",
                -15: "درگاه پرداخت تعلیق شده",
                -16: "سطح تایید Merchant نامعتبر",
                -50: "مبلغ کمتر از حد مجاز",
                -51: "مبلغ بیشتر از حد مجاز",
                -54: "Authority منقضی شده"
            }
        
            error_msg = error_messages.get(error_code, "خطای نامشخص")
        
            await query.message.edit_text(
                f"❌ پرداخت ناموفق بود!\n\n"
                f"کد خطا: {error_code}\n"
                f"توضیح: {error_msg}\n\n"
                f"لطفاً مجدداً تلاش کنید یا با پشتیبانی تماس بگیرید."
            )



    # ==================== MY SERVICES ====================
    elif data == "my_services":
        orders = get_user_orders(user_id)
        active_orders = [o for o in orders if o['status'] == 'active']
        
        if not active_orders:
            text = "❌ شما هیچ سرویس فعالی ندارید.\n\nبرای خرید سرویس جدید از منوی اصلی استفاده کنید."
            keyboard = [[InlineKeyboardButton("🛒 خرید سرویس", callback_data="buy_service")],
                       [InlineKeyboardButton("🏠 منوی اصلی", callback_data="back_to_main")]]
            await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            text = "📊 <b>سرویس‌های فعال شما:</b>\n\n"
            keyboard = []
            
            for order in active_orders[:10]:
                pkg = PACKAGES.get(order['package_id'], {})
                text += f"🔹 {pkg.get('name', 'نامشخص')}\n"
                text += f"   📅 انقضا: {format_date(order['expires_at'])}\n\n"
                
                keyboard.append([InlineKeyboardButton(
                    f"📊 {pkg.get('name', 'نامشخص')}", 
                    callback_data=f"service_detail_{order['id']}"
                )])
            
            keyboard.append([InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")])
            await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data.startswith("service_detail_"):
        order_id = int(data[15:])
        orders = get_user_orders(user_id)
        order = next((o for o in orders if o['id'] == order_id), None)
        
        if not order:
            await query.answer("❌ سرویس یافت نشد", show_alert=True)
            return
        
        # دریافت اطلاعات از Marzban
        usage = await marzban.get_user_usage(order['marzban_username'])
        
        pkg = PACKAGES.get(order['package_id'], {})
        text = f"📊 <b>جزئیات سرویس</b>\n\n"
        text += f"📦 پکیج: {pkg.get('name', 'نامشخص')}\n"
        text += f"👤 نام کاربری: <code>{order['marzban_username']}</code>\n"
        
        if usage:
            text += f"📊 مصرف شده: {usage['used_gb']} GB از {usage['total_gb']} GB\n"
            text += f"📊 باقیمانده: {usage['remaining_gb']} GB\n"
            text += f"📅 تاریخ انقضا: {format_date(usage['expire'])}\n"
            text += f"🔌 وضعیت: {'✅ فعال' if usage['status'] == 'active' else '❌ غیرفعال'}\n"
        
        text += f"\n🔗 لینک اتصال:\n<code>{order['subscription_url']}</code>"
        
        keyboard = [
            [InlineKeyboardButton("🔄 تمدید سرویس", callback_data=f"renew_{order_id}")],
            [InlineKeyboardButton("🗑 حذف سرویس", callback_data=f"delete_service_{order_id}")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data="my_services")]
        ]
        
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data.startswith("renew_"):
        order_id = int(data[6:])
        # TODO: پیاده‌سازی تمدید
        await query.answer("⚠️ تمدید سرویس به زودی فعال می‌شود", show_alert=True)

    elif data.startswith("delete_service_"):
        order_id = int(data[15:])
        keyboard = [
            [InlineKeyboardButton("✅ بله، حذف شود", callback_data=f"confirm_delete_{order_id}")],
            [InlineKeyboardButton("❌ خیر", callback_data=f"service_detail_{order_id}")]
        ]
        await safe_edit_message(query, "⚠️ آیا مطمئن هستید که می‌خواهید این سرویس را حذف کنید؟", reply_markup=InlineKeyboardMarkup(keyboard))

    elif data.startswith("confirm_delete_"):
        order_id = int(data[15:])
        orders = get_user_orders(user_id)
        order = next((o for o in orders if o['id'] == order_id), None)

        if order:
            success = await marzban.delete_user(order['marzban_username'])
        
            if success:
                update_order_status(order_id, 'deleted')
                await query.message.edit_text("✅ سرویس با موفقیت حذف شد.")

            # 🏠 نمایش منوی اصلی بعد از حذف
                text = "🏠از گزینه‌های زیر یکی رو انتخاب کن /n ✅ سرویس با موفقیت حذف شد.👇"
                keyboard = [[InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")]]
                await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
            else:
                await query.message.edit_text("❌ خطا در حذف سرویس. لطفاً با پشتیبانی تماس بگیرید.")
        else:
            await query.message.edit_text("❌ سفارش مورد نظر پیدا نشد یا قبلاً حذف شده است.")

            

    # ==================== USER PROFILE ====================
    elif data == "user_profile":
        orders = get_user_orders(user_id)
        active_count = len([o for o in orders if o['status'] == 'active'])
        
        text = f"👤 <b>پروفایل کاربری</b>\n\n"
        text += f"🆔 شناسه: <code>{user_id}</code>\n"
        text += f"👤 نام: {db_user['first_name']}\n"
        if db_user['username']:
            text += f"🔗 نام کاربری: @{db_user['username']}\n"
        text += f"💰 موجودی: {format_price(db_user['balance'])}\n"
        text += f"📊 سرویس‌های فعال: {active_count}\n"
        text += f"💵 کل خریدها: {format_price(db_user['total_purchased'])}\n"
        text += f"📅 عضویت: {format_date(db_user['created_at'])}\n"
        
        keyboard = [[InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")]]
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    # ==================== WALLET ====================
    elif data == "wallet":
        transactions = get_transactions(user_id, 10)
        
        text = f"💰 <b>کیف پول</b>\n\n"
        text += f"💵 موجودی فعلی: {format_price(db_user['balance'])}\n\n"
        text += "<b>آخرین تراکنش‌ها:</b>\n"
        
        for t in transactions:
            sign = "+" if t['amount'] > 0 else ""
            text += f"• {sign}{format_price(t['amount'])} - {t['description'][:30]}\n"
            text += f"  {format_date(t['created_at'])}\n"
        
        keyboard = [
            [InlineKeyboardButton("➕ افزایش موجودی", callback_data="charge_wallet")],
            [InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")]
        ]
        
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data == "charge_wallet":
        text = "💰 <b>افزایش موجودی کیف پول</b>\n\n"
        text += "لطفاً مبلغ مورد نظر را به تومان وارد کنید:\n\n"
        text += "مثال: 50000"
        
        keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="wallet")]]
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
        
        context.user_data['state'] = WAITING_WALLET_CHARGE_AMOUNT

    # ==================== REFERRAL ====================
    elif data == "referral":
        referral_code = db_user['referral_code']
        bot_username = context.bot.username
        referral_link = f"https://t.me/{bot_username}?start={referral_code}"
        
        invited_count = get_invited_users_count(user_id)
        inviter_reward = int(get_setting('referral_inviter_reward', '10000'))
        total_earned = invited_count * inviter_reward
        
        text = f"🎁 <b>دعوت دوستان</b>\n\n"
        text += f"🔗 لینک اختصاصی شما:\n<code>{referral_link}</code>\n\n"
        text += f"💰 پاداش هر دعوت: {format_price(inviter_reward)}\n"
        text += f"👥 تعداد دعوت‌شدگان: {invited_count} نفر\n"
        text += f"💵 کل درآمد از دعوت: {format_price(total_earned)}\n\n"
        text += "دوستان خود را دعوت کنید و برای هر نفر پاداش دریافت کنید! 🎉"
        
        keyboard = [[InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")]]
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    # ==================== HELP ====================
    elif data == "help":
        text = "❓ <b>راهنمای استفاده</b>\n\n"
        text += "1️⃣ از منوی «خرید سرویس» پکیج مورد نظر را انتخاب کنید\n"
        text += "2️⃣ با کیف پول یا پرداخت آنلاین خریداری کنید\n"
        text += "3️⃣ لینک اتصال را کپی کرده و در اپلیکیشن V2Ray وارد کنید\n\n"
        text += "📱 اپلیکیشن‌های پیشنهادی:\n"
        text += "• Android: v2rayNG\n"
        text += "• iOS: Fair VPN, Shadowrocket\n"
        text += "• Windows: v2rayN\n\n"
        text += "برای مشکلات با پشتیبانی تماس بگیرید."
        
        keyboard = [[InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")]]
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    # ==================== SUPPORT ====================
    elif data == "support":
        text = "📞 <b>پشتیبانی</b>\n\n"
        text += "برای ارتباط با پشتیبانی از راه‌های زیر استفاده کنید:\n\n"
        text += "📱 تلگرام: @boleyla1\n"
        text += "📧 ایمیل: boleyla.mehrshad1@gmail.com\n\n"
        text += "⏰ ساعت پاسخگویی: 9 صبح تا 12 شب"
        
        keyboard = [[InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")]]
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    # ==================== ADMIN PANEL ====================
    elif data == "admin_panel":
            await show_admin_panel(query, context)

    elif data == "admin_dashboard":
        await show_admin_dashboard(query)

    elif data == "admin_users":
        await show_admin_users_menu(query, context)

    elif data == "admin_services":
        await show_admin_services_menu(query)

    elif data == "admin_financial":
        await show_admin_financial_menu(query)

    elif data == "admin_referral":
        await show_admin_referral_menu(query)

    elif data == "admin_settings":
        await show_admin_settings_menu(query)
     
    elif data == "admin_services_active":
        await show_admin_services_list(query, status='active')
    
    elif data == "admin_services_expired":
        await show_admin_services_list(query, status='expired')
    
    elif data == "admin_services_stats":
        await show_admin_services_stats(query)
    
    elif data.startswith("admin_services_") and "_page_" in data:
        # مثال: admin_services_active_page_1
        parts = data.split("_")
        status = parts[2]
        page = int(parts[4])
        await show_admin_services_list(query, status=status, page=page)
    
    elif data.startswith("admin_service_detail_"):
        order_id = int(data.split("_")[-1])
        await show_admin_service_detail(query, order_id, context)
    
    elif data == "admin_services_search":
        context.user_data['state'] = WAITING_SERVICE_SEARCH
        await query.message.edit_text(
            "🔍 لطفاً نام کاربری Marzban یا شناسه سفارش را وارد کنید:"
        )
    
    # افزایش حجم
    elif data.startswith("admin_service_addtraffic_"):
        order_id = int(data.split("_")[-1])
        context.user_data['state'] = WAITING_TRAFFIC_AMOUNT
        context.user_data['target_order_id'] = order_id
        
        keyboard = [
            [InlineKeyboardButton("10 GB", callback_data=f"admin_addtraffic_{order_id}_10")],
            [InlineKeyboardButton("20 GB", callback_data=f"admin_addtraffic_{order_id}_20")],
            [InlineKeyboardButton("50 GB", callback_data=f"admin_addtraffic_{order_id}_50")],
            [InlineKeyboardButton("100 GB", callback_data=f"admin_addtraffic_{order_id}_100")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data=f"admin_service_detail_{order_id}")]
        ]
        
        await safe_edit_message(
            query,
            "📦 چه مقدار حجم اضافه شود؟\n\nیا عدد دلخواه (به GB) ارسال کنید:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    elif data.startswith("admin_addtraffic_"):
        # مثال: admin_addtraffic_123_20
        parts = data.split("_")
        order_id = int(parts[2])
        gb_amount = int(parts[3])
        
        await process_add_traffic(query, order_id, gb_amount, context)
    
    # تمدید سرویس
    elif data.startswith("admin_service_extend_"):
        order_id = int(data.split("_")[-1])
        context.user_data['state'] = WAITING_EXTEND_DAYS
        context.user_data['target_order_id'] = order_id
        
        keyboard = [
            [InlineKeyboardButton("7 روز", callback_data=f"admin_extend_{order_id}_7")],
            [InlineKeyboardButton("15 روز", callback_data=f"admin_extend_{order_id}_15")],
            [InlineKeyboardButton("30 روز", callback_data=f"admin_extend_{order_id}_30")],
            [InlineKeyboardButton("60 روز", callback_data=f"admin_extend_{order_id}_60")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data=f"admin_service_detail_{order_id}")]
        ]
        
        await safe_edit_message(
            query,
            "⏰ چند روز تمدید شود؟\n\nیا تعداد روز دلخواه را ارسال کنید:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    elif data.startswith("admin_extend_"):
        # مثال: admin_extend_123_30
        parts = data.split("_")
        order_id = int(parts[2])
        days = int(parts[3])
        
        await process_extend_service(query, order_id, days, context)
    
    # غیرفعال کردن
    elif data.startswith("admin_service_disable_"):
        order_id = int(data.split("_")[-1])
        
        keyboard = [
            [InlineKeyboardButton("✅ بله، غیرفعال شود", callback_data=f"admin_confirm_disable_{order_id}")],
            [InlineKeyboardButton("❌ خیر", callback_data=f"admin_service_detail_{order_id}")]
        ]
        
        await safe_edit_message(
            query,
            "⚠️ آیا مطمئن هستید که می‌خواهید این سرویس را غیرفعال کنید؟",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    elif data.startswith("admin_confirm_disable_"):
        order_id = int(data.split("_")[-1])
        await process_disable_service(query, order_id, context)
    
    # حذف سرویس
    elif data.startswith("admin_service_delete_"):
        order_id = int(data.split("_")[-1])
        
        keyboard = [
            [InlineKeyboardButton("✅ بله، حذف شود", callback_data=f"admin_confirm_servdel_{order_id}")],
            [InlineKeyboardButton("❌ خیر", callback_data=f"admin_service_detail_{order_id}")]
        ]
        
        await safe_edit_message(
            query,
            "⚠️ <b>هشدار!</b>\n\nآیا مطمئن هستید؟ این عمل قابل بازگشت نیست!",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )
    
    elif data.startswith("admin_confirm_servdel_"):
        order_id = int(data.split("_")[-1])
        await process_delete_service_admin(query, order_id, context)
    
    # فعال‌سازی مجدد
    elif data.startswith("admin_service_reactivate_"):
        order_id = int(data.split("_")[-1])
        await process_reactivate_service(query, order_id, context)
        # ================== بازگشت‌ها ==================
    elif data.startswith("admin_back_"):
        section = data.replace("admin_back_", "")
        if section == "panel":
            await show_admin_panel(query, context)
        elif section == "users":
            await show_admin_users_menu(query, context)
        elif section == "services":
                await show_admin_services_menu(query)
        elif section == "financial":
                await show_admin_financial_menu(query)
        elif section == "referral":
                await show_admin_referral_menu(query)
        elif section == "settings":
            await show_admin_settings_menu(query)
            
            

        # ================== مدیریت کاربران ==================
    elif data == "admin_user_search":
        context.user_data['state'] = WAITING_USER_SEARCH
        await query.message.edit_text("🔍 لطفاً نام یا آیدی کاربر مورد نظر را ارسال کنید:")

    elif data == "admin_user_list":
        users = get_all_users(limit=10)
        text = "📋 <b>لیست کاربران اخیر:</b>\n\n"
        keyboard = []
        for u in users:
            text += f"👤 {u['first_name']} (@{u['username'] or 'بدون نام'})\n🆔 {u['user_id']}\n\n"
            keyboard.append([InlineKeyboardButton(
                f"👤 {u['first_name']} - {u['user_id']}",
                callback_data=f"admin_view_user_{u['user_id']}"
            )])
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_users")])
        await safe_edit_message(query, text, InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data.startswith("admin_view_user_"):
        user_id = int(data.split("_")[-1])
        user = get_user(user_id)
        if not user:
            await query.message.edit_text("❌ کاربر یافت نشد.")
            return
        text = f"👤 <b>{user['first_name']}</b>\n"
        text += f"🆔 {user['user_id']}\n"
        text += f"💰 موجودی: {format_price(user['balance'])}\n"
        text += f"📅 تاریخ عضویت: {user['created_at']}\n"
        keyboard = [
            [InlineKeyboardButton("💰 تغییر موجودی", callback_data=f"admin_edit_balance_{user_id}")],
            [InlineKeyboardButton("🗑 حذف کاربر", callback_data=f"admin_delete_user_{user_id}")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_users")]
        ]
        await safe_edit_message(query, text, InlineKeyboardMarkup(keyboard), parse_mode='HTML')

        # ================== سایر ==================
    elif data == "admin_broadcast":
        context.user_data['state'] = WAITING_BROADCAST_MESSAGE
        await query.message.edit_text("📢 لطفاً پیام مورد نظر را ارسال کنید تا برای همه کاربران ارسال شود:")
        # ==================== ADMIN USER TAGS & BULK BALANCE ====================
    
    elif data == "admin_bulk_balance":
        await show_admin_bulk_balance_menu(query, context)
    
    elif data.startswith("admin_bulkbal_tag_"):
        tag = data.replace("admin_bulkbal_tag_", "")
        context.user_data['bulk_balance_tag'] = tag
        context.user_data['state'] = WAITING_BULK_BALANCE_AMOUNT
        
        user_count = get_users_count(tag=tag if tag != 'all' else None)
        
        await query.message.edit_text(
            f"💰 <b>افزایش موجودی گروهی</b>\n\n"
            f"🏷 تگ: <code>{tag}</code>\n"
            f"👥 تعداد کاربران: {user_count}\n\n"
            f"💵 مبلغ افزایش را به تومان وارد کنید:",
            parse_mode='HTML'
        )
    
    elif data == "admin_user_tags":
        await show_admin_user_tags_stats(query)
    
    elif data == "admin_create_tag":
        await show_create_tag_menu(query, context)
    
    elif data == "admin_edit_tags":
        await show_edit_tags_menu(query)
    
    elif data.startswith("admin_edittag_"):
        tag = data.replace("admin_edittag_", "")
        await show_tag_edit_options(query, tag)
    
    elif data.startswith("admin_renametag_"):
        tag = data.replace("admin_renametag_", "")
        context.user_data['state'] = WAITING_USER_TAG
        context.user_data['renaming_tag'] = tag
        
        await query.message.edit_text(
            f"✏️ نام جدید برای تگ <code>{tag}</code> را وارد کنید:",
            parse_mode='HTML'
        )
    
    elif data.startswith("admin_deletetag_"):
        tag = data.replace("admin_deletetag_", "")
        keyboard = [
            [InlineKeyboardButton("✅ بله، حذف شود", callback_data=f"admin_confirmdeltag_{tag}")],
            [InlineKeyboardButton("❌ خیر", callback_data=f"admin_edittag_{tag}")]
        ]
        
        user_count = get_users_count(tag=tag)
        
        await safe_edit_message(
            query,
            f"⚠️ <b>هشدار!</b>\n\n"
            f"آیا مطمئن هستید که می‌خواهید تگ <code>{tag}</code> را حذف کنید؟\n\n"
            f"👥 {user_count} کاربر به تگ <code>regular</code> تبدیل می‌شوند.",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )
    
    elif data.startswith("admin_confirmdeltag_"):
        tag = data.replace("admin_confirmdeltag_", "")
        await process_delete_tag(query, tag)



    # در callback_query_handler اضافه کنید:

    elif data == "admin_search_coupon":
        await start_search_coupon(query, context)

    elif data == "admin_coupon_stats":
        await show_coupon_stats(query)

    elif data.startswith("admin_coupon_detail_"):
        coupon_id = int(data.split("_")[-1])
        await show_coupon_detail(query, coupon_id)

    elif data.startswith("admin_coupon_fullstats_"):
        coupon_id = int(data.split("_")[-1])
        await show_coupon_full_stats(query, coupon_id)

    elif data.startswith("admin_coupon_enable_"):
        coupon_id = int(data.split("_")[-1])
        await toggle_coupon_status(query, coupon_id, enable=True)

    elif data.startswith("admin_coupon_disable_"):
        coupon_id = int(data.split("_")[-1])
        await toggle_coupon_status(query, coupon_id, enable=False)

    elif data.startswith("admin_coupon_delete_"):
        coupon_id = int(data.split("_")[-1])
    
    # تایید حذف
        keyboard = [
            [
                InlineKeyboardButton("✅ بله، حذف شود", callback_data=f"admin_coupon_delete_confirm_{coupon_id}"),
                InlineKeyboardButton("❌ انصراف", callback_data=f"admin_coupon_detail_{coupon_id}")
            ]
        ]

        await query.message.edit_text(
            "⚠️ <b>تایید حذف کوپن</b>\n\n"
            "آیا مطمئن هستید؟ این عملیات غیرقابل بازگشت است!",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    elif data.startswith("admin_coupon_delete_confirm_"):
        coupon_id = int(data.split("_")[-1])
        await delete_coupon(query, coupon_id)




    # ==================== ADMIN EDIT BALANCE ====================
    elif data.startswith("admin_edit_balance_"):
        target_user_id = int(data.split("_")[-1])
        target_user = get_user(target_user_id)
    
        if not target_user:
            await query.message.edit_text("❌ کاربر یافت نشد.")
            return
    
        context.user_data['state'] = WAITING_BALANCE_AMOUNT
        context.user_data['target_user_id'] = target_user_id
    
        text = f"💰 <b>تغییر موجودی</b>\n\n"
        text += f"👤 کاربر: {target_user['first_name']}\n"
        text += f"💵 موجودی فعلی: {format_price(target_user['balance'])}\n\n"
        text += "لطفاً مبلغ تغییر را وارد کنید:\n\n"
        text += "• عدد مثبت: افزایش موجودی (مثال: 50000)\n"
        text += "• عدد منفی: کاهش موجودی (مثال: -20000)"
    
        keyboard = [[InlineKeyboardButton("🔙 انصراف", callback_data=f"admin_view_user_{target_user_id}")]]
    
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

# ==================== ADMIN DELETE USER ====================
    elif data.startswith("admin_delete_user_"):
        target_user_id = int(data.split("_")[-1])
        target_user = get_user(target_user_id)
    
        if not target_user:
            await query.message.edit_text("❌ کاربر یافت نشد.")
            return
    
    # شمارش سرویس‌های فعال کاربر
        active_orders = get_user_orders(target_user_id)
        active_count = len([o for o in active_orders if o['status'] == 'active'])
    
        text = f"⚠️ <b>حذف کاربر</b>\n\n"
        text += f"👤 نام: {target_user['first_name']}\n"
        text += f"🆔 آیدی: <code>{target_user_id}</code>\n"
        text += f"💰 موجودی: {format_price(target_user['balance'])}\n"
        text += f"📊 سرویس‌های فعال: {active_count}\n\n"
    
        if active_count > 0:
            text += "⚠️ <b>هشدار:</b> این کاربر سرویس فعال دارد!\n\n"
    
        text += "❌ <b>این عمل قابل بازگشت نیست!</b>\n\n"
        text += "آیا مطمئن هستید؟"
    
        keyboard = [
            [InlineKeyboardButton("✅ بله، حذف شود", callback_data=f"admin_confirm_delete_user_{target_user_id}")],
            [InlineKeyboardButton("❌ خیر", callback_data=f"admin_view_user_{target_user_id}")]
        ]
    
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

# ==================== ADMIN CONFIRM DELETE USER ====================
    elif data.startswith("admin_confirm_delete_user_"):
        target_user_id = int(data.split("_")[-1])
        target_user = get_user(target_user_id)
    
        if not target_user:
            await query.message.edit_text("❌ کاربر یافت نشد.")
            return

        try:
        # حذف سرویس‌های کاربر از Marzban
            orders = get_user_orders(target_user_id)
            deleted_services = 0
        
            for order in orders:
                if order['status'] == 'active':
                    marzban_username = order.get('marzban_username')
                    if marzban_username:
                        success = await marzban.delete_user(marzban_username)
                        if success:
                            deleted_services += 1
        
        # حذف از دیتابیس (CASCADE خودکار orders و transactions را حذف می‌کند)
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM users WHERE user_id = %s", (target_user_id,))
            conn.commit()
            cursor.close()
            conn.close()
        
        # ثبت لاگ
            log_admin_action(
                query.from_user.id,
                'delete_user',
                target_user_id,
                f"حذف کاربر {target_user['first_name']} با {deleted_services} سرویس فعال"
                )
        
            text = f"✅ <b>کاربر حذف شد</b>\n\n"
            text += f"👤 {target_user['first_name']}\n"
            text += f"🆔 {target_user_id}\n"
            text += f"🗑 سرویس‌های حذف شده: {deleted_services}"
        
            keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_users")]]
        
            await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
        
        except Exception as e:
            logger.error(f"خطا در حذف کاربر {target_user_id}: {e}")
            await query.message.edit_text(
                f"❌ خطا در حذف کاربر!\n\n"
                f"لطفاً با پشتیبانی تماس بگیرید.\n\n"
                f"خطا: {str(e)[:100]}"
        )
        else:
            await query.message.reply_text(f"❓ دکمه ناشناخته: {data}")
        # ==================== ADMIN TOP REFERRERS ====================
    elif data == "admin_top_referrers":
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT 
                u.user_id,
                u.first_name,
                u.username,
                COUNT(r.user_id) AS referral_count,
                u.total_purchased
            FROM users u
            LEFT JOIN users r ON r.referred_by = u.user_id
            GROUP BY u.user_id
            HAVING referral_count > 0
            ORDER BY referral_count DESC
            LIMIT 10
        """)

        top_referrers = cursor.fetchall()
        cursor.close()
        conn.close()

        if not top_referrers:
            text = "📊 هنوز هیچ دعوتی ثبت نشده است."
            keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_referral")]]
        else:
            text = "🏆 <b>برترین دعوت‌کنندگان</b>\n\n"
            inviter_reward = int(get_setting('referral_inviter_reward', '10000'))

            for idx, ref in enumerate(top_referrers, 1):
                medal = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉" if idx == 3 else f"{idx}️⃣"
                username_text = f"@{ref['username']}" if ref['username'] else "بدون نام کاربری"

                text += f"{medal} <b>{ref['first_name']}</b> ({username_text})\n"
                text += f"   👥 دعوت‌ها: {ref['referral_count']} نفر\n"
                text += f"   💰 درآمد رفرال: {format_price(ref['referral_count'] * inviter_reward)}\n"
                text += f"   💳 کل خرید: {format_price(ref['total_purchased'] or 0)}\n\n"

            keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_referral")]]

        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


# ==================== ADMIN SET INVITER REWARD ====================
    elif data == "admin_set_inviter_reward":
        current = get_setting('referral_inviter_reward', '10000')
    
        context.user_data['state'] = WAITING_REFERRAL_REWARD_INVITER
    
        text = f"⚙️ <b>تنظیم پاداش دعوت‌کننده</b>\n\n"
        text += f"💰 مبلغ فعلی: {format_price(int(current))}\n\n"
        text += "مبلغ جدید را به تومان وارد کنید:"
    
        keyboard = [[InlineKeyboardButton("🔙 انصراف", callback_data="admin_referral")]]
    
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

# ==================== ADMIN SET INVITED REWARD ====================
    elif data == "admin_set_invited_reward":
        current = get_setting('referral_invited_reward', '5000')
    
        context.user_data['state'] = WAITING_REFERRAL_REWARD_INVITED
    
        text = f"⚙️ <b>تنظیم پاداش دعوت‌شده</b>\n\n"
        text += f"💰 مبلغ فعلی: {format_price(int(current))}\n\n"
        text += "مبلغ جدید را به تومان وارد کنید:"
    
        keyboard = [[InlineKeyboardButton("🔙 انصراف", callback_data="admin_referral")]]
    
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    
    # در تابع callback_query_handler اضافه کنید:

    elif data == "admin_financial":
        await show_admin_financial_menu(query)

    elif data == "admin_transactions":
        await show_admin_transactions(query)

    elif data.startswith("admin_transactions_page_"):
        page = int(data.split("_")[-1])
        await show_admin_transactions(query, page)

    elif data == "admin_payments":
        await show_admin_payments(query)

    elif data.startswith("admin_payments_page_"):
        page = int(data.split("_")[-1])
        await show_admin_payments(query, page)

    elif data == "admin_coupons":
        await show_admin_coupons_menu(query)

    elif data == "admin_create_coupon":
        await start_create_coupon(query, context)

    elif data == "admin_list_coupons":
        await show_admin_coupons_list(query)

    elif data.startswith("admin_coupons_filter_"):
        filter_type = data.split("_")[-1]
        await show_admin_coupons_list(query, page=1, filter_type=filter_type)

    elif data.startswith("admin_coupons_list_"):
        parts = data.split("_")
        filter_type = parts[3]
        page = int(parts[4])
        await show_admin_coupons_list(query, page, filter_type)

    elif data.startswith("coupon_type_"):
        coupon_type = data.split("_")[-1]
        context.user_data['coupon_data']['discount_type'] = coupon_type
    
        type_text = "درصد تخفیف" if coupon_type == "percent" else "مبلغ تخفیف"
        await query.message.edit_text(
            f"✅ نوع: {type_text}\n\n"
            f"مقدار تخفیف را وارد کنید:\n"
            f"{'(عدد بین 1 تا 100)' if coupon_type == 'percent' else '(به تومان)'}"
        )

    elif data == "admin_campaigns":
        await show_admin_campaigns_menu(query)

    elif data == "admin_export_excel":
        await export_financial_excel(query, context, 'all')

    elif data == "admin_export_transactions":
        await export_financial_excel(query, context, 'transactions')

    elif data == "admin_export_payments":
        await export_financial_excel(query, context, 'payments')

    elif data == "admin_export_coupon_stats":
        await export_coupon_stats_excel(query, context)

    elif data.startswith("admin_export_coupon_"):
        coupon_id = int(data.split("_")[-1])
        await export_coupon_stats_excel(query, context, coupon_id)
        # ==================== ADMIN SETTINGS HANDLERS ====================

    elif data == "admin_settings":
        await show_admin_settings_menu(query)

    elif data == "admin_marzban_settings":
        current_url = get_setting('marzban_url', MARZBAN_URL)
        current_username = get_setting('marzban_username', MARZBAN_USERNAME)
    
        text = "🔧 <b>تنظیمات Marzban</b>\n\n"
        text += f"🌐 URL: <code>{current_url}</code>\n"
        text += f"👤 Username: <code>{current_username}</code>\n"
        text += f"🔑 Password: ●●●●●●\n\n"
        text += "برای تغییر هر مورد، از دکمه‌های زیر استفاده کنید:"
    
        keyboard = [
            [InlineKeyboardButton("🌐 تغییر URL", callback_data="admin_change_marzban_url")],
            [InlineKeyboardButton("👤 تغییر Username", callback_data="admin_change_marzban_user")],
            [InlineKeyboardButton("🔑 تغییر Password", callback_data="admin_change_marzban_pass")],
            [InlineKeyboardButton("🔌 تست اتصال", callback_data="admin_test_marzban")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_settings")]
        ]
    
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data == "admin_change_marzban_url":
        context.user_data['state'] = WAITING_MARZBAN_URL
    
        current = get_setting('marzban_url', MARZBAN_URL)
    
        await query.message.edit_text(
            f"🌐 <b>تغییر URL مرزبان</b>\n\n"
            f"URL فعلی: <code>{current}</code>\n\n"
            f"URL جدید را وارد کنید:\n"
            f"مثال: <code>https://panel.example.com:8000</code>",
            parse_mode='HTML'
        )

    elif data == "admin_change_marzban_user":
        context.user_data['state'] = WAITING_MARZBAN_USER
    
        current = get_setting('marzban_username', MARZBAN_USERNAME)
    
        await query.message.edit_text(
            f"👤 <b>تغییر نام کاربری مرزبان</b>\n\n"
            f"نام کاربری فعلی: <code>{current}</code>\n\n"
            f"نام کاربری جدید را وارد کنید:",
            parse_mode='HTML'
        )

    elif data == "admin_change_marzban_pass":
        context.user_data['state'] = WAITING_MARZBAN_PASS
    
        await query.message.edit_text(
            f"🔑 <b>تغییر رمز عبور مرزبان</b>\n\n"
            f"رمز عبور جدید را وارد کنید:\n\n"
            f"⚠️ توجه: رمز عبور به صورت امن ذخیره می‌شود.",
            parse_mode='HTML'
        )

    elif data == "admin_test_marzban":
        await query.message.edit_text("⏳ در حال تست اتصال به مرزبان...")
    
    # بروزرسانی تنظیمات marzban از دیتابیس
        marzban_url = get_setting('marzban_url', MARZBAN_URL)
        marzban_username = get_setting('marzban_username', MARZBAN_USERNAME)
        marzban_password = get_setting('marzban_password', MARZBAN_PASSWORD)
    
    # ایجاد نمونه جدید برای تست
        test_marzban = MarzbanAPI(marzban_url, marzban_username, marzban_password)
    
        success = await test_marzban.test_connection()
    
        if success:
            text = "✅ <b>اتصال موفق!</b>\n\n"
            text += f"🌐 URL: <code>{marzban_url}</code>\n"
            text += f"👤 کاربر: <code>{marzban_username}</code>\n"
            text += f"🔌 وضعیت: متصل\n\n"
            text += f"🕐 زمان تست: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"
        else:
            text = "❌ <b>خطا در اتصال!</b>\n\n"
            text += f"🌐 URL: <code>{marzban_url}</code>\n"
            text += f"👤 کاربر: <code>{marzban_username}</code>\n\n"
            text += "⚠️ لطفاً تنظیمات را بررسی کنید."
    
        keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_marzban_settings")]]
    
        await query.message.edit_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data == "admin_zarinpal_settings":
        current_merchant = get_setting('zarinpal_merchant', ZARINPAL_MERCHANT)
    
        text = "💳 <b>تنظیمات زرین‌پال</b>\n\n"
        text += f"🔑 Merchant ID: <code>{current_merchant}</code>\n"
        text += f"🧪 حالت آزمایشی: {'✅ فعال' if ZARINPAL_SANDBOX else '❌ غیرفعال'}\n\n"
        text += "برای تغییر Merchant ID از دکمه زیر استفاده کنید:"
    
        keyboard = [
            [InlineKeyboardButton("✏️ تغییر Merchant ID", callback_data="admin_change_merchant")],
            [InlineKeyboardButton("🧪 تست درگاه", callback_data="admin_test_zarinpal")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_settings")]
        ]
    
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data == "admin_change_merchant":
        context.user_data['state'] = WAITING_MERCHANT_ID
    
        current = get_setting('zarinpal_merchant', ZARINPAL_MERCHANT)
    
        await query.message.edit_text(
            f"💳 <b>تغییر Merchant ID زرین‌پال</b>\n\n"
            f"Merchant فعلی:\n<code>{current}</code>\n\n"
            f"Merchant ID جدید را وارد کنید:\n"
            f"مثال: <code>xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx</code>",
            parse_mode='HTML'
        )

    elif data == "admin_test_zarinpal":
        await query.message.edit_text("⏳ در حال تست درگاه زرین‌پال...")
    
        merchant_id = get_setting('zarinpal_merchant', ZARINPAL_MERCHANT)
        zp = ZarinPal(merchant_id, ZARINPAL_SANDBOX)
    
    # تست با مبلغ 1000 تومان
        result = zp.request_payment(
            amount=1000,
            description="تست اتصال",
            callback_url="https://example.com/verify"
        )
    
        if result.get('data', {}).get('code') == 100:
            text = "✅ <b>اتصال به زرین‌پال موفق!</b>\n\n"
            text += f"🔑 Merchant ID: <code>{merchant_id[:20]}...</code>\n"
            text += f"🧪 حالت: {'Sandbox' if ZARINPAL_SANDBOX else 'Production'}\n"
            text += f"✅ وضعیت: فعال"
        else:
            text = "❌ <b>خطا در اتصال به زرین‌پال!</b>\n\n"
            text += f"🔑 Merchant ID: <code>{merchant_id[:20]}...</code>\n"
            text += f"❌ کد خطا: {result.get('data', {}).get('code', 'نامشخص')}\n\n"
            text += "⚠️ لطفاً Merchant ID را بررسی کنید."
    
        keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_zarinpal_settings")]]
    
        await query.message.edit_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data == "admin_welcome_message":
        context.user_data['state'] = WAITING_WELCOME_MESSAGE

    # پیام فعلی
        current = get_setting(
            'welcome_message', 
            'سلام {user_name} عزیز! 👋\nبه ربات VPN ما خوش آمدید.\nموجودی شما: {balance}\n🎁 هدیه {invited_reward} از {referrer_name}'
        )

        text = "📝 <b>ویرایش پیام خوش‌آمدگویی</b>\n\n"
        text += "━━━━━━━━━━━━━━━━\n\n"
        text += f"<b>پیام فعلی:</b>\n{current}\n\n"
        text += "━━━━━━━━━━━━━━━━\n\n"
        text += "<b>📌 متغیرهای پایه:</b>\n"
        text += "• <code>{user_name}</code> → نام کاربر جدید\n"
        text += "• <code>{user_id}</code> → آیدی عددی\n"
        text += "• <code>{balance}</code> → موجودی کیف پول\n\n"
        text += "<b>🎁 متغیرهای رفرال:</b>\n"
        text += "• <code>{referrer_name}</code> → نام دعوت‌کننده\n"
        text += "• <code>{referrer_username}</code> → یوزرنیم دعوت‌کننده\n"
        text += "• <code>{invited_reward}</code> → مبلغ هدیه\n"
        text += "• <code>{inviter_reward}</code> → پاداش دعوت‌کننده\n\n"
        text += "━━━━━━━━━━━━━━━━\n\n"
        text += "<b>💡 نکته:</b> متغیرهای رفرال فقط برای کاربران جدید نمایش داده می‌شوند.\n\n"
        text += "━━━━━━━━━━━━━━━━\n\n"
        text += "<b>مثال پیش‌نمایش:</b>\n"

    # نمونه کاربر و رفرال برای پیش‌نمایش
        sample_user = {"first_name": "محمد", "id": 123456789}
        sample_referrer = {"first_name": "علی", "username": "ali123"}
        sample_invited_reward = 5000
        sample_inviter_reward = 10000

        def safe_preview(text, user, db_user, referrer=None, invited_reward=0, inviter_reward=0):
            import re

            replacements = {
                'user_name': user.get('first_name', 'کاربر'),
                'first_name': user.get('first_name', 'کاربر'),
                'user_id': str(user.get('id', 0)),
                'balance': format_price(db_user.get('balance', 50000)),  # پیش‌فرض موجودی
                'referrer_name': referrer.get('first_name', 'یک دوست') if referrer else '',
                'referrer_username': f"@{referrer['username']}" if referrer and referrer.get('username') else '',
                'invited_reward': format_price(invited_reward) if referrer else '',
                'inviter_reward': format_price(inviter_reward) if referrer else ''
            }

            def replace_var(match):
                var_name = match.group(1)
                return str(replacements.get(var_name, match.group(0)))

            result = re.sub(r'\{(\w+)\}', replace_var, text)

        # اگر رفرال نیست، بخش هدیه را پاک کنیم
            if not referrer:
                result = re.sub(r'🎁 هدیه .*? از .*?\n?', '', result)

            return result

        preview_with_referral = safe_preview(current, sample_user, {"balance": 50000},
                                         sample_referrer, sample_invited_reward, sample_inviter_reward)
        preview_without_referral = safe_preview(current, sample_user, {"balance": 50000})

        text += f"<code>📌 با لینک دعوت:</code>\n{preview_with_referral}\n\n"
        text += f"<code>📌 بدون لینک دعوت:</code>\n{preview_without_referral}\n\n"
        text += "━━━━━━━━━━━━━━━━\n\n"
        text += "✏️ پیام جدید را ارسال کنید:"

        keyboard = [[InlineKeyboardButton("🔙 انصراف", callback_data="admin_settings")]]

        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')



    elif data == "admin_logs":
        logs = get_admin_logs(limit=20)
    
        if not logs:
            text = "📜 <b>لاگ‌های سیستم</b>\n\n❌ هیچ لاگی یافت نشد."
        else:
            text = "📜 <b>آخرین فعالیت‌های ادمین</b>\n\n"
        
            for log in logs[:15]:
                admin_id = log.get('admin_id', 0)
                action = log.get('action', 'نامشخص')
                details = log.get('details', '')[:50]
                created = format_date(log['created_at'])
            
                text += f"🔹 <b>{action}</b>\n"
                text += f"   👤 ادمین: <code>{admin_id}</code>\n"
                text += f"   📝 {details}\n"
                text += f"   🕐 {created}\n\n"
    
        keyboard = [
            [InlineKeyboardButton("🗑 پاک کردن لاگ‌ها", callback_data="admin_clear_logs")],
            [InlineKeyboardButton("📥 دانلود لاگ", callback_data="admin_export_logs")],
            [InlineKeyboardButton("🔄 بروزرسانی", callback_data="admin_logs")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_settings")]
        ]
    
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

    elif data == "admin_clear_logs":
        keyboard = [
        [InlineKeyboardButton("✅ بله، پاک شود", callback_data="admin_confirm_clear_logs")],
        [InlineKeyboardButton("❌ خیر", callback_data="admin_logs")]
    ]
    
        await safe_edit_message(
            query,
            "⚠️ <b>حذف لاگ‌ها</b>\n\nآیا مطمئن هستید که می‌خواهید تمام لاگ‌ها را پاک کنید؟",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    elif data == "admin_confirm_clear_logs":
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM admin_logs")
        deleted_count = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()
    
        log_admin_action(query.from_user.id, 'clear_logs', None, f"{deleted_count} لاگ حذف شد")
    
        await query.answer(f"✅ {deleted_count} لاگ حذف شد", show_alert=True)
        await show_admin_settings_menu(query)

    elif data == "admin_export_logs":
        try:
            logs = get_admin_logs(limit=1000)
        
            if not logs:
                await query.answer("❌ هیچ لاگی برای خروجی وجود ندارد", show_alert=True)
                return
        
        # ایجاد فایل متنی
            text_content = "📜 گزارش لاگ‌های سیستم\n"
            text_content += f"📅 تاریخ: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}\n"
            text_content += "="*50 + "\n\n"
        
            for log in logs:
                text_content += f"ID: {log['id']}\n"
                text_content += f"ادمین: {log['admin_id']}\n"
                text_content += f"عملیات: {log['action']}\n"
                text_content += f"جزئیات: {log.get('details', 'ندارد')}\n"
                text_content += f"تاریخ: {format_date(log['created_at'])}\n"
                text_content += "-"*50 + "\n"
        
        # تبدیل به بایت
            log_file = io.BytesIO(text_content.encode('utf-8'))
            log_file.name = f"admin_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=log_file,
                filename=log_file.name,
                caption="📜 فایل لاگ‌های سیستم"
            )
        
            await query.answer("✅ فایل ارسال شد", show_alert=False)
        
        except Exception as e:
            logger.error(f"خطا در export logs: {e}")
            await query.answer("❌ خطا در ایجاد فایل", show_alert=True)
    # در تابع button_handler، قبل از else آخر اضافه کنید:

    elif data == "insufficient_balance":
        await query.answer(
            "💰 موجودی کیف پول شما کافی نیست!\n\n"
            "لطفاً ابتدا کیف پول خود را شارژ کنید یا از پرداخت آنلاین استفاده کنید.",
            show_alert=True
        )
    
    elif data == "charge_online":
        amount = context.user_data.get('charge_amount')
    
        if not amount:
            await query.answer("❌ خطا: مبلغ مشخص نشده", show_alert=True)
            return
    
        await query.message.edit_text("⏳ در حال انتقال به درگاه پرداخت...")
    
        merchant_id = get_setting('zarinpal_merchant', ZARINPAL_MERCHANT)
        zp = ZarinPal(merchant_id, ZARINPAL_SANDBOX)
    
        result = zp.request_payment(
            amount=amount * 10,  # تبدیل تومان به ریال ✅
            description=f"شارژ کیف پول",
            mobile=db_user.get('phone'),
            callback_url="https://bot.boleyla.com/zarinpal/callback"  # ✅ درست

        )
    
        if result.get('data', {}).get('code') == 100:
            authority = result['data']['authority']
            payment_url = zp.get_payment_url(authority)
        
        # ذخیره
            save_payment(
                user_id=user_id,
                amount=amount * 10,
                authority=authority,
                package_id=None,
                payment_type='wallet'
            )
        
            text = f"💳 <b>پرداخت آنلاین</b>\n\n"
            text += f"💰 مبلغ شارژ: {format_price(amount)}\n\n"
            text += f"🔗 لطفاً روی دکمه زیر کلیک کنید و پرداخت را انجام دهید.\n\n"
            text += f"✅ بعد از پرداخت موفق، ربات خودکار پیام خواهد فرستاد."
        
            keyboard = [
                [InlineKeyboardButton("💳 پرداخت", url=payment_url)],
                [InlineKeyboardButton("🔙 انصراف", callback_data="wallet")]
            ]
        
            await query.message.edit_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode='HTML'
            )
        
            context.user_data.pop('charge_amount', None)
        else:
            error_code = result.get('data', {}).get('code', 'نامشخص')
            await query.message.edit_text(
                f"❌ خطا در اتصال به درگاه پرداخت!\n\n"
                f"کد خطا: {error_code}\n\n"
                f"لطفاً با پشتیبانی تماس بگیرید."
        )   


# ==================== ADMIN PANEL FUNCTIONS ====================

async def show_admin_panel(query, context):
    stats = get_stats()
    
    text = "⚙️ <b>پنل مدیریت</b>\n\n"
    text += f"👥 کاربران: {stats['total_users']} (امروز: +{stats['new_users_today']})\n"
    text += f"📊 سرویس‌های فعال: {stats['active_orders']}\n"
    text += f"💰 درآمد امروز: {format_price(stats['today_sales'])}\n"
    text += f"💵 درآمد کل: {format_price(stats['total_income'])}\n"
    
    keyboard = [
        [InlineKeyboardButton("📊 داشبورد", callback_data="admin_dashboard")],
        [InlineKeyboardButton("👥 مدیریت کاربران", callback_data="admin_users"),
         InlineKeyboardButton("📦 مدیریت سرویس‌ها", callback_data="admin_services")],
        [InlineKeyboardButton("💰 مدیریت مالی", callback_data="admin_financial"),
         InlineKeyboardButton("🎁 مدیریت رفرال", callback_data="admin_referral")],
        [InlineKeyboardButton("📢 ارسال پیام همگانی", callback_data="admin_broadcast")],
        [InlineKeyboardButton("⚙️ تنظیمات", callback_data="admin_settings")],
        [InlineKeyboardButton("🏠 بازگشت", callback_data="back_to_main")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

async def show_admin_dashboard(query):
    stats = get_stats()
    
    text = "📊 <b>داشبورد مدیریت</b>\n\n"
    text += "<b>📈 آمار کاربران:</b>\n"
    text += f"• کل کاربران: {stats['total_users']}\n"
    text += f"• کاربران جدید امروز: {stats['new_users_today']}\n\n"
    
    text += "<b>💰 آمار مالی:</b>\n"
    text += f"• فروش امروز: {format_price(stats['today_sales'])}\n"
    text += f"• فروش این هفته: {format_price(stats['week_sales'])}\n"
    text += f"• فروش این ماه: {format_price(stats['month_sales'])}\n"
    text += f"• کل درآمد: {format_price(stats['total_income'])}\n"
    text += f"• موجودی کل کیف پول‌ها: {format_price(stats['total_balance'])}\n\n"
    
    text += "<b>📊 سرویس‌ها:</b>\n"
    text += f"• سرویس‌های فعال: {stats['active_orders']}\n"
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_panel")]]
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

async def show_admin_users_menu(query, context):
    text = "👥 <b>مدیریت کاربران</b>\n\nگزینه مورد نظر را انتخاب کنید:"
    
    keyboard = [
        [InlineKeyboardButton("🔍 جستجوی کاربر", callback_data="admin_user_search")],
        [InlineKeyboardButton("📋 لیست کاربران", callback_data="admin_user_list")],
        [InlineKeyboardButton("💰 افزایش موجودی گروهی", callback_data="admin_bulk_balance")],
        [InlineKeyboardButton("📊 آمار تگ‌ها", callback_data="admin_user_tags")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_panel")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

# ==================== ADMIN SERVICES MANAGEMENT ====================

async def show_admin_services_menu(query):
    """منوی اصلی مدیریت سرویس‌ها"""
    stats = get_stats()
    
    # آمار سرویس‌ها
    all_orders = get_all_orders()
    active = len([o for o in all_orders if o['status'] == 'active'])
    expired = len([o for o in all_orders if o['status'] == 'expired'])
    deleted = len([o for o in all_orders if o['status'] == 'deleted'])
    
    text = "📦 <b>مدیریت سرویس‌ها</b>\n\n"
    text += f"✅ سرویس‌های فعال: {active}\n"
    text += f"⏰ منقضی شده: {expired}\n"
    text += f"🗑 حذف شده: {deleted}\n"
    text += f"📊 کل سرویس‌ها: {len(all_orders)}\n\n"
    text += "یک گزینه را انتخاب کنید:"

    keyboard = [
        [InlineKeyboardButton("📋 لیست سرویس‌های فعال", callback_data="admin_services_active")],
        [InlineKeyboardButton("🔍 جستجوی سرویس", callback_data="admin_services_search")],
        [InlineKeyboardButton("⏰ سرویس‌های منقضی شده", callback_data="admin_services_expired")],
        [InlineKeyboardButton("📊 آمار کامل", callback_data="admin_services_stats")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_panel")]
    ]

    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_admin_services_list(query, status='active', page=0):
    """نمایش لیست سرویس‌ها با صفحه‌بندی"""
    per_page = 8
    offset = page * per_page
    
    orders = get_all_orders(status=status, limit=per_page + 1)
    
    if not orders:
        text = f"❌ هیچ سرویس {status} یافت نشد."
        keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_services")]]
        await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard))
        return
    
    has_more = len(orders) > per_page
    orders = orders[:per_page]
    
    status_emoji = {
        'active': '✅',
        'expired': '⏰',
        'deleted': '🗑'
    }
    
    text = f"{status_emoji.get(status, '📦')} <b>سرویس‌های {status}</b>\n\n"
    
    keyboard = []
    for order in orders:
        username = order.get('marzban_username', 'N/A')
        user_name = order.get('first_name', 'ناشناس')
        
        # دریافت نام پکیج از package_id
        pkg = PACKAGES.get(order.get('package_id', ''), {})
        package_name = pkg.get('name', 'نامشخص')
        
        # محاسبه روزهای باقی‌مانده
        expire = order.get('expires_at')  # نام صحیح فیلد
        days_left = ""
        if expire:
            try:
                if isinstance(expire, str):
                    expire_dt = datetime.fromisoformat(expire)
                else:
                    expire_dt = expire
                days = (expire_dt - datetime.now()).days
                if days > 0:
                    days_left = f" ({days} روز)"
                else:
                    days_left = " (منقضی)"
            except:
                pass
        
        button_text = f"👤 {user_name} | {package_name}{days_left}"
        keyboard.append([InlineKeyboardButton(
            button_text,
            callback_data=f"admin_service_detail_{order['id']}"
        )])
    
    # دکمه‌های صفحه‌بندی
    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton("◀️ قبلی", callback_data=f"admin_services_{status}_page_{page-1}"))
    if has_more:
        nav_buttons.append(InlineKeyboardButton("بعدی ▶️", callback_data=f"admin_services_{status}_page_{page+1}"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_services")])
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')





async def show_admin_service_detail(query, order_id, context):
    """نمایش جزئیات کامل یک سرویس"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    # دریافت اطلاعات سفارش
    cursor.execute("""
        SELECT o.*, u.first_name, u.username as user_username, u.user_id
        FROM orders o
        JOIN users u ON o.user_id = u.user_id
        WHERE o.id = %s
    """, (order_id,))
    order = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if not order:
        await query.message.edit_text("❌ سرویس یافت نشد.")
        return
    
    # دریافت نام پکیج
    pkg = PACKAGES.get(order.get('package_id', ''), {})
    package_name = pkg.get('name', 'نا مشخص')
    
    # دریافت اطلاعات از مرزبان
    marzban_username = order.get('marzban_username')
    usage_info = None
    
    if marzban_username:
        usage_info = await marzban.get_user_usage(marzban_username)
    
    # ساخت متن
    text = "📦 <b>جزئیات سرویس</b>\n\n"
    text += f"🆔 شناسه سفارش: <code>{order['id']}</code>\n"
    text += f"👤 کاربر: {order['first_name']} (@{order.get('user_username', 'ندارد')})\n"
    text += f"🔢 آیدی کاربر: <code>{order['user_id']}</code>\n"
    text += f"📦 پکیج: {package_name}\n"
    text += f"💰 قیمت: {format_price(order['price'])}\n"
    text += f"📅 تاریخ خرید: {format_date(order['created_at'])}\n"
    
    # بررسی فیلد expires_at
    if order.get('expires_at'):
        text += f"⏰ تاریخ انقضا: {format_date(order['expires_at'])}\n"
        try:
            expire_dt = order['expires_at']
            if isinstance(expire_dt, str):
                expire_dt = datetime.fromisoformat(expire_dt)
            days_left = (expire_dt - datetime.now()).days
            if days_left > 0:
                text += f"⏳ روزهای باقیمانده: {days_left} روز\n"
            else:
                text += f"❌ منقضی شده: {abs(days_left)} روز پیش\n"
        except Exception as e:
            logger.error(f"خطا در محاسبه روزهای باقیمانده: {e}")
    
    status_emoji = {
        'active': '✅ فعال',
        'expired': '⏰ منقضی شده',
        'deleted': '🗑 حذف شده'
    }
    text += f"📊 وضعیت: {status_emoji.get(order['status'], order['status'])}\n"
    
    # اطلاعات مرزبان
    if usage_info:
        text += f"\n📈 <b>آمار مصرف:</b>\n"
        text += f"💾 مصرف شده: {usage_info['used_gb']} GB\n"
        text += f"📦 کل حجم: {usage_info['total_gb']} GB\n"
        text += f"✅ باقیمانده: {usage_info['remaining_gb']} GB\n"
        
        if usage_info['total'] > 0:
            usage_percent = (usage_info['used'] / usage_info['total']) * 100
            text += f"📊 درصد مصرف: {usage_percent:.1f}%\n"
        
        if usage_info.get('subscription_url'):
            text += f"\n🔗 <b>لینک اتصال:</b>\n<code>{usage_info['subscription_url']}</code>\n"
    
    # دکمه‌های عملیات
    keyboard = []
    
    if order['status'] == 'active':
        keyboard.append([
            InlineKeyboardButton("➕ افزایش حجم", callback_data=f"admin_service_addtraffic_{order_id}"),
            InlineKeyboardButton("⏰ تمدید", callback_data=f"admin_service_extend_{order_id}")
        ])
        keyboard.append([
            InlineKeyboardButton("⏸ غیرفعال کردن", callback_data=f"admin_service_disable_{order_id}"),
            InlineKeyboardButton("🗑 حذف", callback_data=f"admin_service_delete_{order_id}")
        ])
    elif order['status'] == 'expired':
        keyboard.append([
            InlineKeyboardButton("♻️ فعال‌سازی مجدد", callback_data=f"admin_service_reactivate_{order_id}"),
            InlineKeyboardButton("🗑 حذف", callback_data=f"admin_service_delete_{order_id}")
        ])
    
    keyboard.append([
        InlineKeyboardButton("👤 مشاهده کاربر", callback_data=f"admin_view_user_{order['user_id']}"),
        InlineKeyboardButton("🔄 بروزرسانی", callback_data=f"admin_service_detail_{order_id}")
    ])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_services_active")])
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')



async def show_admin_services_stats(query):
    """نمایش آمار کامل سرویس‌ها"""
    all_orders = get_all_orders()
    
    total = len(all_orders)
    active = len([o for o in all_orders if o['status'] == 'active'])
    expired = len([o for o in all_orders if o['status'] == 'expired'])
    deleted = len([o for o in all_orders if o['status'] == 'deleted'])
    
    # محاسبه درآمد
    total_revenue = sum(o['price'] for o in all_orders if o['status'] in ['active', 'expired'])
    
    # سرویس‌های امروز
    today = datetime.now().date()
    today_orders = [o for o in all_orders if o.get('created_at') and o['created_at'].date() == today]
    
    # سرویس‌های در حال انقضا (کمتر از 7 روز)
    expiring_soon = []
    for order in all_orders:
        if order['status'] == 'active' and order.get('expires_at'):  # اصلاح نام فیلد
            try:
                expire_dt = order['expires_at']
                if isinstance(expire_dt, str):
                    expire_dt = datetime.fromisoformat(expire_dt)
                days = (expire_dt - datetime.now()).days
                if 0 < days <= 7:
                    expiring_soon.append(order)
            except Exception as e:
                logger.error(f"خطا در محاسبه انقضا: {e}")
                continue
    
    # پرفروش‌ترین پکیج‌ها
    from collections import Counter
    package_ids = [o.get('package_id') for o in all_orders if o.get('package_id')]
    package_counter = Counter(package_ids)
    top_packages = []
    
    for pkg_id, count in package_counter.most_common(3):
        pkg = PACKAGES.get(pkg_id, {})
        pkg_name = pkg.get('name', pkg_id)
        top_packages.append((pkg_name, count))
    
    text = "📊 <b>آمار کامل سرویس‌ها</b>\n\n"
    text += "<b>📈 وضعیت سرویس‌ها:</b>\n"
    text += f"• کل سرویس‌ها: {total}\n"
    text += f"• ✅ فعال: {active} ({(active/total*100 if total > 0 else 0):.1f}%)\n"
    text += f"• ⏰ منقضی: {expired} ({(expired/total*100 if total > 0 else 0):.1f}%)\n"
    text += f"• 🗑 حذف شده: {deleted}\n\n"
    
    text += f"<b>💰 درآمد:</b>\n"
    text += f"• کل درآمد: {format_price(total_revenue)}\n"
    text += f"• میانگین فروش: {format_price(total_revenue // total if total > 0 else 0)}\n\n"
    
    text += f"<b>📅 امروز:</b>\n"
    text += f"• سرویس‌های جدید: {len(today_orders)}\n\n"
    
    text += f"<b>⚠️ در حال انقضا:</b>\n"
    text += f"• سرویس‌های کمتر از 7 روز: {len(expiring_soon)}\n\n"
    
    if top_packages:
        text += "<b>🏆 پرفروش‌ترین پکیج‌ها:</b>\n"
        for pkg, count in top_packages:
            text += f"• {pkg}: {count} عدد\n"
    
    keyboard = [
        [InlineKeyboardButton("⚠️ سرویس‌های در حال انقضا", callback_data="admin_services_expiring")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_services")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

# ==================== ADMIN USER MANAGEMENT - ADDITIONAL FEATURES ====================

async def show_admin_bulk_balance_menu(query, context):
    """منوی افزایش موجودی گروهی"""
    text = "💰 <b>افزایش موجودی گروهی</b>\n\n"
    text += "می‌توانید به گروهی از کاربران بر اساس تگ، موجودی اضافه کنید.\n\n"
    text += "مراحل:\n"
    text += "1️⃣ انتخاب تگ کاربران\n"
    text += "2️⃣ وارد کردن مبلغ\n"
    text += "3️⃣ تایید نهایی"
    
    # دریافت تگ‌های موجود
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT user_tag FROM users WHERE user_tag IS NOT NULL AND user_tag != ''")
    tags = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    
    keyboard = []
    
    if tags:
        for tag in tags:
            # شمارش کاربران با این تگ
            user_count = get_users_count(tag=tag)
            keyboard.append([
                InlineKeyboardButton(
                    f"🏷 {tag} ({user_count} کاربر)",
                    callback_data=f"admin_bulkbal_tag_{tag}"
                )
            ])
    else:
        text += "\n\n⚠️ هیچ تگی تعریف نشده است!"
    
    keyboard.append([InlineKeyboardButton("➕ همه کاربران", callback_data="admin_bulkbal_tag_all")])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_users")])
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_admin_user_tags_stats(query):
    """نمایش آمار تگ‌های کاربران"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    # دریافت آمار هر تگ
    cursor.execute("""
        SELECT 
            user_tag,
            COUNT(*) as user_count,
            SUM(balance) as total_balance,
            SUM(total_purchased) as total_purchased
        FROM users
        WHERE user_tag IS NOT NULL AND user_tag != ''
        GROUP BY user_tag
        ORDER BY user_count DESC
    """)
    
    tag_stats = cursor.fetchall()
    
    # کاربران بدون تگ
    cursor.execute("""
        SELECT COUNT(*) as count, SUM(balance) as balance, SUM(total_purchased) as purchased
        FROM users 
        WHERE user_tag IS NULL OR user_tag = ''
    """)
    no_tag = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    text = "🏷 <b>آمار تگ‌های کاربران</b>\n\n"
    
    if tag_stats:
        for tag in tag_stats:
            text += f"<b>🏷 {tag['user_tag']}</b>\n"
            text += f"  👥 کاربران: {tag['user_count']}\n"
            text += f"  💰 موجودی کل: {format_price(tag['total_balance'] or 0)}\n"
            text += f"  💳 خرید کل: {format_price(tag['total_purchased'] or 0)}\n\n"
    
    if no_tag and no_tag['count'] > 0:
        text += f"<b>❓ بدون تگ</b>\n"
        text += f"  👥 کاربران: {no_tag['count']}\n"
        text += f"  💰 موجودی کل: {format_price(no_tag['balance'] or 0)}\n"
        text += f"  💳 خرید کل: {format_price(no_tag['purchased'] or 0)}\n"
    
    keyboard = [
        [InlineKeyboardButton("➕ ایجاد تگ جدید", callback_data="admin_create_tag")],
        [InlineKeyboardButton("✏️ ویرایش تگ‌ها", callback_data="admin_edit_tags")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_users")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def process_bulk_balance_addition(query, tag, amount, reason, context):
    """اعمال افزایش موجودی گروهی"""
    # دریافت کاربران بر اساس تگ
    if tag == 'all':
        users = get_all_users(limit=10000)  # همه کاربران
    else:
        users = get_all_users(tag=tag, limit=10000)
    
    if not users:
        await query.message.edit_text("❌ هیچ کاربری یافت نشد!")
        return
    
    success_count = 0
    fail_count = 0
    
    conn = db.get_connection()
    cursor = conn.cursor()
    
    for user in users:
        try:
            # افزایش موجودی
            cursor.execute(
                "UPDATE users SET balance = balance + %s WHERE user_id = %s",
                (amount, user['user_id'])
            )
            
            # ثبت تراکنش
            cursor.execute(
                """INSERT INTO transactions 
                (user_id, amount, type, description, admin_id) 
                VALUES (%s, %s, %s, %s, %s)""",
                (user['user_id'], amount, 'admin_add', 
                 f"افزایش گروهی - {reason}", query.from_user.id)
            )
            
            conn.commit()
            success_count += 1
            
            # ارسال اطلاعیه به کاربر
            try:
                await context.bot.send_message(
                    user['user_id'],
                    f"🎁 <b>شارژ هدیه!</b>\n\n"
                    f"به حساب شما {format_price(amount)} اضافه شد! 🎉\n\n"
                    f"📝 دلیل: {reason}\n"
                    f"💰 موجودی جدید: {format_price(user['balance'] + amount)}",
                    parse_mode='HTML'
                )
            except Exception as e:
                logger.error(f"خطا در ارسال پیام به {user['user_id']}: {e}")
        
        except Exception as e:
            logger.error(f"خطا در افزایش موجودی {user['user_id']}: {e}")
            fail_count += 1
            conn.rollback()
    
    cursor.close()
    conn.close()
    
    # لاگ ادمین
    log_admin_action(
        query.from_user.id, 
        'bulk_balance',
        0,
        f"افزایش {format_price(amount)} به {success_count} کاربر با تگ {tag}"
    )
    
    text = f"✅ <b>افزایش موجودی گروهی انجام شد!</b>\n\n"
    text += f"👥 کاربران موفق: {success_count}\n"
    text += f"❌ خطا: {fail_count}\n"
    text += f"💰 مبلغ هر کاربر: {format_price(amount)}\n"
    text += f"💵 مجموع: {format_price(amount * success_count)}\n"
    text += f"📝 دلیل: {reason}"
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_users")]]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_create_tag_menu(query, context):
    """منوی ایجاد تگ جدید"""
    text = "➕ <b>ایجاد تگ جدید</b>\n\n"
    text += "تگ‌های پیش‌فرض:\n"
    text += "• <code>regular</code> - کاربران عادی\n"
    text += "• <code>vip</code> - کاربران ویژه\n"
    text += "• <code>premium</code> - کاربران پرمیوم\n"
    text += "• <code>bronze</code> - برنزی\n"
    text += "• <code>silver</code> - نقره‌ای\n"
    text += "• <code>gold</code> - طلایی\n\n"
    text += "یا می‌توانید تگ دلخواه خود را بسازید.\n\n"
    text += "🏷 نام تگ جدید را ارسال کنید:"
    
    context.user_data['state'] = WAITING_USER_TAG
    context.user_data['creating_new_tag'] = True
    
    keyboard = [[InlineKeyboardButton("🔙 انصراف", callback_data="admin_user_tags")]]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_edit_tags_menu(query):
    """منوی ویرایش تگ‌ها"""
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT user_tag, COUNT(*) as count 
        FROM users 
        WHERE user_tag IS NOT NULL AND user_tag != ''
        GROUP BY user_tag
    """)
    tags = cursor.fetchall()
    cursor.close()
    conn.close()
    
    text = "✏️ <b>ویرایش تگ‌ها</b>\n\n"
    text += "یک تگ را برای تغییر نام یا حذف انتخاب کنید:\n\n"
    
    keyboard = []
    
    for tag, count in tags:
        text += f"• {tag} ({count} کاربر)\n"
        keyboard.append([
            InlineKeyboardButton(
                f"🏷 {tag}",
                callback_data=f"admin_edittag_{tag}"
            )
        ])
    
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_user_tags")])
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_tag_edit_options(query, tag):
    """نمایش گزینه‌های ویرایش یک تگ"""
    user_count = get_users_count(tag=tag)
    
    text = f"✏️ <b>ویرایش تگ: {tag}</b>\n\n"
    text += f"👥 تعداد کاربران: {user_count}\n\n"
    text += "چه کاری می‌خواهید انجام دهید؟"
    
    keyboard = [
        [InlineKeyboardButton("✏️ تغییر نام", callback_data=f"admin_renametag_{tag}")],
        [InlineKeyboardButton("🗑 حذف تگ", callback_data=f"admin_deletetag_{tag}")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_edit_tags")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def process_delete_tag(query, tag):
    """حذف یک تگ (تبدیل به regular)"""
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # تبدیل همه کاربران این تگ به regular
    cursor.execute(
        "UPDATE users SET user_tag = 'regular' WHERE user_tag = %s",
        (tag,)
    )
    affected = cursor.rowcount
    conn.commit()
    cursor.close()
    conn.close()
    
    log_admin_action(query.from_user.id, 'delete_tag', 0, f"حذف تگ {tag} - {affected} کاربر")
    
    text = f"✅ تگ <code>{tag}</code> حذف شد.\n\n"
    text += f"👥 {affected} کاربر به تگ <code>regular</code> تبدیل شدند."
    
    keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data="admin_user_tags")]]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

# ==================== HELPER FUNCTIONS (FIXED) ====================

async def process_add_traffic(query, order_id, gb_amount, context):
    """افزایش حجم سرویس"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM orders WHERE id = %s", (order_id,))
    order = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if not order:
        await query.message.edit_text("❌ سفارش یافت نشد.")
        return
    
    marzban_username = order.get('marzban_username')
    if not marzban_username:
        await query.message.edit_text("❌ نام کاربری مرزبان یافت نشد.")
        return
    
    # دریافت اطلاعات فعلی
    user_data = await marzban.get_user(marzban_username)
    if not user_data:
        await query.message.edit_text("❌ کاربر در مرزبان یافت نشد.")
        return
    
    current_limit = user_data.get('data_limit', 0)
    new_limit = current_limit + (gb_amount * 1024**3)  # تبدیل به بایت
    
    # بروزرسانی در مرزبان
    success = await marzban.modify_user(
        username=marzban_username,
        data_limit=new_limit
    )
    
    if success:
        # دریافت نام پکیج
        pkg = PACKAGES.get(order.get('package_id', ''), {})
        package_name = pkg.get('name', 'نامشخص')
        
        await query.message.edit_text(
            f"✅ <b>حجم اضافه شد!</b>\n\n"
            f"• حجم اضافه شده: {gb_amount} GB\n"
            f"• حجم قبلی: {current_limit / 1024**3:.2f} GB\n"
            f"• حجم جدید: {new_limit / 1024**3:.2f} GB",
            parse_mode='HTML'
        )
        
        # لاگ ادمین
        log_admin_action(query.from_user.id, 'add_traffic', order['user_id'], 
                        f"افزایش {gb_amount}GB به سرویس {marzban_username}")
        
        # ارسال اطلاعیه به کاربر
        try:
            await context.bot.send_message(
                order['user_id'],
                f"✨ <b>افزایش حجم سرویس</b>\n\n"
                f"سرویس شما {gb_amount} GB حجم اضافه دریافت کرد! 🎉\n\n"
                f"📦 پکیج: {package_name}\n"
                f"💾 حجم جدید: {new_limit / 1024**3:.2f} GB",
                parse_mode='HTML'
            )
        except Exception as e:
            logger.error(f"خطا در ارسال اطلاعیه: {e}")
    else:
        await query.message.edit_text("❌ خطا در افزایش حجم.")




# ==================== HELPER FUNCTIONS ====================

async def process_add_traffic(query, order_id, gb_amount, context):
    """افزایش حجم سرویس"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM orders WHERE id = %s", (order_id,))
    order = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if not order:
        await query.message.edit_text("❌ سفارش یافت نشد.")
        return
    
    marzban_username = order.get('marzban_username')
    if not marzban_username:
        await query.message.edit_text("❌ نام کاربری مرزبان یافت نشد.")
        return
    
    # دریافت اطلاعات فعلی
    user_data = await marzban.get_user(marzban_username)
    if not user_data:
        await query.message.edit_text("❌ کاربر در مرزبان یافت نشد.")
        return
    
    current_limit = user_data.get('data_limit', 0)
    new_limit = current_limit + (gb_amount * 1024**3)  # تبدیل به بایت
    
    # بروزرسانی در مرزبان
    success = await marzban.modify_user(
        username=marzban_username,
        data_limit=new_limit
    )
    
    if success:
        await query.message.edit_text(
            f"✅ <b>حجم اضافه شد!</b>\n\n"
            f"• حجم اضافه شده: {gb_amount} GB\n"
            f"• حجم قبلی: {current_limit / 1024**3:.2f} GB\n"
            f"• حجم جدید: {new_limit / 1024**3:.2f} GB",
            parse_mode='HTML'
        )
        
        # ارسال اطلاعیه به کاربر
        try:
            await context.bot.send_message(
                order['user_id'],
                f"✨ <b>افزایش حجم سرویس</b>\n\n"
                f"سرویس شما {gb_amount} GB حجم اضافه دریافت کرد! 🎉\n\n"
                f"📦 پکیج: {order['package_name']}\n"
                f"💾 حجم جدید: {new_limit / 1024**3:.2f} GB",
                parse_mode='HTML'
            )
        except:
            pass
    else:
        await query.message.edit_text("❌ خطا در افزایش حجم.")


async def process_extend_service(query, order_id, days, context):
    """تمدید سرویس - نسخه اصلاح شده"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM orders WHERE id = %s", (order_id,))
    order = cursor.fetchone()

    if not order:
        cursor.close()
        conn.close()
        await query.message.edit_text("❌ سفارش یافت نشد.")
        return

    marzban_username = order.get('marzban_username')
    if not marzban_username:
        cursor.close()
        conn.close()
        await query.message.edit_text("❌ نام کاربری مرزبان یافت نشد.")
        return

    # ✅ دریافت اطلاعات کامل از Marzban
    user_data = await marzban.get_user(marzban_username)
    if not user_data:
        cursor.close()
        conn.close()
        await query.message.edit_text("❌ کاربر در مرزبان یافت نشد.")
        return

    # ✅ محاسبه تاریخ انقضای جدید
    current_expire_timestamp = user_data.get('expire', 0)

    # تبدیل timestamp
    if current_expire_timestamp:
        if current_expire_timestamp > 10000000000:  # میلی‌ثانیه
            current_expire = datetime.fromtimestamp(current_expire_timestamp / 1000)
        else:  # ثانیه
            current_expire = datetime.fromtimestamp(current_expire_timestamp)
    else:
        current_expire = datetime.now()

    # محاسبه تاریخ جدید
    if current_expire < datetime.now():
        new_expire = datetime.now() + timedelta(days=days)
    else:
        new_expire = current_expire + timedelta(days=days)

    # تبدیل به timestamp ثانیه
    new_expire_timestamp = int(new_expire.timestamp())

    # ✅ بروزرسانی توکن
    if not marzban.token:
        await marzban.get_token()

    try:
        headers = {
            "Authorization": f"Bearer {marzban.token}",
            "Content-Type": "application/json"
        }

        # ✅ ایجاد payload کامل با تمام فیلدهای ضروری
        update_payload = {
            "proxies": user_data.get('proxies', {}),
            "inbounds": user_data.get('inbounds', {}),
            "expire": new_expire_timestamp,
            "data_limit": user_data.get('data_limit', 0),
            "data_limit_reset_strategy": user_data.get('data_limit_reset_strategy', 'no_reset'),
            "status": "active"
        }

        logger.info(f"🔄 تمدید {marzban_username}: expire={new_expire_timestamp}, date={new_expire}")
        logger.info(f"📦 Payload: {update_payload}")

        # ✅ ارسال با PUT
        async with aiohttp.ClientSession() as session:
            async with session.put(
                f"{marzban.url}/api/user/{marzban_username}",
                json=update_payload,
                headers=headers,
                timeout=aiohttp.ClientTimeout(total=15)
            ) as resp:
                response_text = await resp.text()
                
                logger.info(f"📡 Response Status: {resp.status}")
                logger.info(f"📡 Response Body: {response_text}")

                if resp.status == 200:
                    # ✅ موفقیت
                    cursor.execute(
                        "UPDATE orders SET expires_at = %s, status = 'active' WHERE id = %s",
                        (new_expire, order_id)
                    )
                    conn.commit()

                    pkg = PACKAGES.get(order.get('package_id', ''), {})
                    package_name = pkg.get('name', 'نامشخص')

                    await query.message.edit_text(
                        f"✅ <b>سرویس تمدید شد!</b>\n\n"
                        f"📦 پکیج: {package_name}\n"
                        f"⏰ تعداد روز: {days} روز\n"
                        f"📅 تاریخ قبلی: {format_date(current_expire)}\n"
                        f"📅 تاریخ جدید: {format_date(new_expire)}\n"
                        f"👤 Username: <code>{marzban_username}</code>",
                        parse_mode='HTML'
                    )

                    log_admin_action(
                        query.from_user.id,
                        'extend_service',
                        order['user_id'],
                        f"تمدید {days} روزه {marzban_username}"
                    )

                    # ارسال اطلاعیه به کاربر
                    try:
                        await context.bot.send_message(
                            order['user_id'],
                            f"✨ <b>تمدید سرویس</b>\n\n"
                            f"سرویس شما {days} روز تمدید شد! 🎉\n\n"
                            f"📦 پکیج: {package_name}\n"
                            f"⏰ تاریخ انقضا: {format_date(new_expire)}\n"
                            f"👤 Username: <code>{marzban_username}</code>",
                            parse_mode='HTML'
                        )
                    except Exception as e:
                        logger.error(f"خطا در ارسال اطلاعیه: {e}")

                else:
                    logger.error(f"❌ خطای Marzban {resp.status}: {response_text}")
                    await query.message.edit_text(
                        f"❌ خطا در تمدید سرویس\n\n"
                        f"کد خطا: {resp.status}\n"
                        f"پیام خطا:\n<code>{response_text[:500]}</code>\n\n"
                        f"لطفاً لاگ سرور مرزبان را بررسی کنید.",
                        parse_mode='HTML'
                    )

    except asyncio.TimeoutError:
        logger.error("❌ Timeout در اتصال به Marzban")
        await query.message.edit_text(
            "❌ خطا: زمان اتصال به مرزبان به پایان رسید.\n\n"
            "لطفاً اتصال به سرور مرزبان را بررسی کنید."
        )
    except Exception as e:
        logger.error(f"❌ خطا در process_extend_service: {e}", exc_info=True)
        await query.message.edit_text(
            f"❌ خطا در اتصال به Marzban\n\n"
            f"جزئیات: <code>{str(e)}</code>",
            parse_mode='HTML'
        )
    finally:
        cursor.close()
        conn.close()




async def process_disable_service(query, order_id, context):
    """غیرفعال کردن سرویس"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM orders WHERE id = %s", (order_id,))
    order = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if not order:
        await query.message.edit_text("❌ سفارش یافت نشد.")
        return
    
    marzban_username = order.get('marzban_username')
    if not marzban_username:
        await query.message.edit_text("❌ نام کاربری مرزبان یافت نشد.")
        return
    
    # دریافت نام پکیج
    pkg = PACKAGES.get(order.get('package_id', ''), {})
    package_name = pkg.get('name', 'نامشخص')
    
    # بررسی توکن مرزبان
    if not marzban.token:
        await marzban.get_token()
    
    try:
        headers = {
            "Authorization": f"Bearer {marzban.token}",
            "Content-Type": "application/json"
        }
        
        # دریافت اطلاعات فعلی کاربر از مرزبان
        user_data = await marzban.get_user(marzban_username)
        if not user_data:
            await query.message.edit_text("❌ کاربر در مرزبان یافت نشد.")
            return
        
        # ✅ ساخت داده کامل برای غیرفعال‌سازی
        update_data = {
            "username": marzban_username,
            "status": "disabled",  # تغییر وضعیت به disabled
            "proxies": user_data.get('proxies', {}),
            "inbounds": user_data.get('inbounds', {}),
            "expire": user_data.get('expire'),
            "data_limit": user_data.get('data_limit'),
            "data_limit_reset_strategy": user_data.get('data_limit_reset_strategy', 'no_reset'),
            "note": user_data.get('note', ''),
            "on_hold_timeout": user_data.get('on_hold_timeout'),
            "on_hold_expire_duration": user_data.get('on_hold_expire_duration')
        }
        
        # ارسال درخواست به Marzban
        async with aiohttp.ClientSession() as session:
            async with session.put(
                f"{marzban.url}/api/user/{marzban_username}",
                json=update_data,
                headers=headers,
                timeout=aiohttp.ClientTimeout(total=15)
            ) as resp:
                if resp.status == 200:
                    # موفقیت‌آمیز بود
                    conn = db.get_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE orders SET status = %s WHERE id = %s",
                        ('expired', order_id)
                    )
                    conn.commit()
                    cursor.close()
                    conn.close()
                    
                    await query.message.edit_text(
                        f"✅ <b>سرویس غیرفعال شد</b>\n\n"
                        f"📦 پکیج: {package_name}\n"
                        f"👤 کاربر: {order['user_id']}\n"
                        f"🔌 وضعیت Marzban: غیرفعال\n"
                        f"👤 Username: <code>{marzban_username}</code>",
                        parse_mode='HTML'
                    )
                    
                    # لاگ ادمین
                    log_admin_action(
                        query.from_user.id, 
                        'disable_service', 
                        order['user_id'], 
                        f"غیرفعال‌سازی {marzban_username}"
                    )
                    
                    # اطلاعیه به کاربر
                    try:
                        await context.bot.send_message(
                            order['user_id'],
                            f"⚠️ <b>غیرفعال‌سازی سرویس</b>\n\n"
                            f"سرویس {package_name} شما توسط ادمین غیرفعال شد.\n\n"
                            f"👤 Username: <code>{marzban_username}</code>\n\n"
                            f"برای اطلاعات بیشتر با پشتیبانی تماس بگیرید.",
                            parse_mode='HTML'
                        )
                    except Exception as e:
                        logger.error(f"خطا در ارسال اطلاعیه: {e}")
                        
                else:
                    error_text = await resp.text()
                    logger.error(f"خطای Marzban: {resp.status} - {error_text}")
                    await query.message.edit_text(
                        f"❌ خطا در غیرفعال‌سازی در Marzban\n\n"
                        f"کد خطا: {resp.status}\n"
                        f"پیام: {error_text[:200]}"
                    )
    
    except Exception as e:
        logger.error(f"خطا در process_disable_service: {e}")
        await query.message.edit_text(
            f"❌ خطا در اتصال به Marzban\n\n{str(e)}"
        )





async def process_delete_service_admin(query, order_id, context):
    """حذف کامل سرویس"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM orders WHERE id = %s", (order_id,))
    order = cursor.fetchone()
    cursor.close()
    conn.close()
    
    if not order:
        await query.message.edit_text("❌ سفارش یافت نشد.")
        return
    
    marzban_username = order.get('marzban_username')
    
    # حذف از مرزبان
    success = await marzban.delete_user(marzban_username)
    
    # دریافت نام پکیج
    pkg = PACKAGES.get(order.get('package_id', ''), {})
    package_name = pkg.get('name', 'نامشخص')
    
    if success:
        update_order_status(order_id, 'deleted')
        
        await query.message.edit_text(
            f"✅ سرویس با موفقیت حذف شد.\n\n"
            f"📦 پکیج: {package_name}\n"
            f"👤 کاربر: {order['user_id']}"
        )
        
        # لاگ ادمین
        log_admin_action(query.from_user.id, 'delete_service', order['user_id'], 
                        f"حذف سرویس {marzban_username}")
    else:
        await query.message.edit_text("❌ خطا در حذف سرویس از مرزبان.")




async def process_reactivate_service(query, order_id, context):
    """فعال‌سازی مجدد سرویس منقضی"""
    keyboard = [
        [InlineKeyboardButton("⏰ تمدید سرویس", callback_data=f"admin_service_extend_{order_id}")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data=f"admin_service_detail_{order_id}")]
    ]
    
    await safe_edit_message(
        query,
        "⏰ برای فعال‌سازی مجدد، ابتدا باید سرویس را تمدید کنید.\n\n"
        "از دکمه زیر استفاده کنید:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def show_admin_financial_menu(query):
    """منوی مدیریت مالی با آمار"""
    # دریافت آمار مالی
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    # کل تراکنش‌ها
    cursor.execute("SELECT COUNT(*) as count, SUM(amount) as total FROM transactions")
    trans_stats = cursor.fetchone()
    
    # پرداخت‌های موفق
    cursor.execute("""
        SELECT COUNT(*) as count, SUM(amount) as total 
        FROM payments 
        WHERE status = 'success'
    """)
    payment_stats = cursor.fetchone()
    
    # کوپن‌های فعال
    cursor.execute("""
        SELECT COUNT(*) as count 
        FROM coupons 
        WHERE is_active = 1 AND (expires_at IS NULL OR expires_at > NOW())
    """)
    coupon_count = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    text = "💰 <b>مدیریت مالی</b>\n\n"
    text += f"📊 کل تراکنش‌ها: {trans_stats['count'] or 0}\n"
    text += f"💵 مبلغ کل: {format_price(trans_stats['total'] or 0)}\n\n"
    text += f"✅ پرداخت‌های موفق: {payment_stats['count'] or 0}\n"
    text += f"💳 درآمد پرداخت: {format_price(payment_stats['total'] or 0)}\n\n"
    text += f"🎟 کوپن‌های فعال: {coupon_count['count'] or 0}\n\n"
    text += "گزینه مورد نظر را انتخاب کنید:"
    
    keyboard = [
        [InlineKeyboardButton("📊 گزارش تراکنش‌ها", callback_data="admin_transactions")],
        [InlineKeyboardButton("💳 گزارش پرداخت‌ها", callback_data="admin_payments")],
        [InlineKeyboardButton("🎟 مدیریت کوپن‌ها", callback_data="admin_coupons")],
        [InlineKeyboardButton("🎉 مدیریت کمپین‌ها", callback_data="admin_campaigns")],
        [InlineKeyboardButton("📥 دانلود گزارش Excel", callback_data="admin_export_excel")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_panel")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

async def show_admin_transactions(query, page=1):
    """نمایش تراکنش‌های مالی"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    per_page = 10
    offset = (page - 1) * per_page
    
    # دریافت تراکنش‌ها
    cursor.execute("""
        SELECT t.*, u.first_name, u.username 
        FROM transactions t
        LEFT JOIN users u ON t.user_id = u.user_id
        ORDER BY t.created_at DESC
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    transactions = cursor.fetchall()
    
    # تعداد کل
    cursor.execute("SELECT COUNT(*) as total FROM transactions")
    total = cursor.fetchone()['total']
    
    cursor.close()
    conn.close()
    
    total_pages = (total + per_page - 1) // per_page
    
    if not transactions:
        text = "📊 <b>گزارش تراکنش‌ها</b>\n\n❌ تراکنشی یافت نشد."
    else:
        text = f"📊 <b>گزارش تراکنش‌ها</b>\n\n"
        text += f"📄 صفحه {page} از {total_pages}\n"
        text += f"📈 کل تراکنش‌ها: {total}\n\n"
        
        for trans in transactions:
            trans_type = {
                'charge': '➕ شارژ',
                'purchase': '🛒 خرید',
                'refund': '↩️ بازگشت',
                'referral': '🎁 رفرال',
                'admin_add': '⚙️ ادمین',
                'admin_deduct': '⚠️ کسر'
            }.get(trans['type'], trans['type'])
            
            name = trans.get('first_name', 'نامشخص')
            username = f"@{trans['username']}" if trans.get('username') else 'بدون یوزر'
            
            text += f"🔹 <b>{trans_type}</b>\n"
            text += f"   👤 {name} ({username})\n"
            text += f"   💰 {format_price(trans['amount'])}\n"
            text += f"   📅 {format_date(trans['created_at'])}\n"
            if trans.get('description'):
                text += f"   📝 {trans['description'][:50]}\n"
            text += "\n"
    
    # دکمه‌های صفحه‌بندی
    keyboard = []
    nav_buttons = []
    
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("◀️ قبلی", callback_data=f"admin_transactions_page_{page-1}"))
    
    nav_buttons.append(InlineKeyboardButton(f"📄 {page}/{total_pages}", callback_data="noop"))
    
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("بعدی ▶️", callback_data=f"admin_transactions_page_{page+1}"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    keyboard.append([
        InlineKeyboardButton("📥 Excel", callback_data="admin_export_transactions"),
        InlineKeyboardButton("🔍 جستجو", callback_data="admin_search_transaction")
    ])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_financial")])
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_admin_payments(query, page=1):
    """نمایش گزارش پرداخت‌های آنلاین"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    per_page = 10
    offset = (page - 1) * per_page
    
    cursor.execute("""
        SELECT p.*, u.first_name, u.username
        FROM payments p
        LEFT JOIN users u ON p.user_id = u.user_id
        ORDER BY p.created_at DESC
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    payments = cursor.fetchall()
    
    cursor.execute("SELECT COUNT(*) as total FROM payments")
    total = cursor.fetchone()['total']
    
    # آمار وضعیت‌ها
    cursor.execute("""
        SELECT 
            status,
            COUNT(*) as count,
            SUM(amount) as total
        FROM payments
        GROUP BY status
    """)
    status_stats = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    total_pages = (total + per_page - 1) // per_page
    
    text = f"💳 <b>گزارش پرداخت‌ها</b>\n\n"
    
    # آمار
    for stat in status_stats:
        status_emoji = {
            'pending': '⏳',
            'success': '✅',
            'failed': '❌',
            'cancelled': '🚫'
        }.get(stat['status'], '❓')
        
        text += f"{status_emoji} {stat['status']}: {stat['count']} ({format_price(stat['total'] or 0)})\n"
    
    text += f"\n📄 صفحه {page} از {total_pages}\n\n"
    
    if payments:
        for payment in payments:
            status_emoji = {
                'pending': '⏳',
                'success': '✅',
                'failed': '❌',
                'cancelled': '🚫'
            }.get(payment['status'], '❓')
            
            name = payment.get('first_name', 'نامشخص')
            username = f"@{payment['username']}" if payment.get('username') else ''
            
            text += f"{status_emoji} <b>{payment['status']}</b>\n"
            text += f"   👤 {name} {username}\n"
            text += f"   💰 {format_price(payment['amount'])}\n"
            text += f"   🔢 Authority: <code>{payment.get('authority', 'N/A')[:20]}</code>\n"
            text += f"   📅 {format_date(payment['created_at'])}\n\n"
    else:
        text += "❌ پرداختی یافت نشد."
    
    # دکمه‌ها
    keyboard = []
    nav_buttons = []
    
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("◀️", callback_data=f"admin_payments_page_{page-1}"))
    nav_buttons.append(InlineKeyboardButton(f"{page}/{total_pages}", callback_data="noop"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("▶️", callback_data=f"admin_payments_page_{page+1}"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    keyboard.append([
        InlineKeyboardButton("📥 Excel", callback_data="admin_export_payments"),
        InlineKeyboardButton("🔄 بروزرسانی", callback_data="admin_payments")
    ])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_financial")])
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_admin_coupons_menu(query):
    """منوی مدیریت کوپن‌ها"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    # آمار کوپن‌ها
    cursor.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN is_active = 1 THEN 1 ELSE 0 END) as active,
            SUM(used_count) as total_used
        FROM coupons
    """)
    stats = cursor.fetchone()
    
    # کوپن‌های اخیر
    cursor.execute("""
        SELECT * FROM coupons 
        ORDER BY created_at DESC 
        LIMIT 5
    """)
    recent_coupons = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    text = "🎟 <b>مدیریت کوپن‌های تخفیف</b>\n\n"
    text += f"📊 کل کوپن‌ها: {stats['total'] or 0}\n"
    text += f"✅ فعال: {stats['active'] or 0}\n"
    text += f"📈 استفاده شده: {stats['total_used'] or 0}\n\n"
    
    if recent_coupons:
        text += "<b>🕐 آخرین کوپن‌ها:</b>\n"
        for coupon in recent_coupons:
            status = "✅" if coupon['is_active'] else "❌"
            discount_type = f"{coupon['discount_percent']}%" if coupon.get('discount_percent') else f"{format_price(coupon.get('discount_amount', 0))}"
            
            text += f"{status} <code>{coupon['code']}</code> - {discount_type}\n"
    
    keyboard = [
        [InlineKeyboardButton("➕ ایجاد کوپن جدید", callback_data="admin_create_coupon")],
        [InlineKeyboardButton("📋 لیست کوپن‌ها", callback_data="admin_list_coupons")],
        [InlineKeyboardButton("🔍 جستجوی کوپن", callback_data="admin_search_coupon")],
        [InlineKeyboardButton("📊 آمار استفاده", callback_data="admin_coupon_stats")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_financial")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def start_create_coupon(query, context):
    """شروع فرآیند ایجاد کوپن"""
    context.user_data['state'] = WAITING_COUPON_DATA
    context.user_data['coupon_data'] = {}
    
    text = "➕ <b>ایجاد کوپن تخفیف</b>\n\n"
    text += "لطفاً کد کوپن را وارد کنید:\n"
    text += "(فقط حروف انگلیسی، اعداد و خط تیره)\n\n"
    text += "مثال: <code>SUMMER2025</code>"
    
    keyboard = [[InlineKeyboardButton("🔙 انصراف", callback_data="admin_coupons")]]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

async def process_coupon_creation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """پردازش مراحل ایجاد کوپن طبق دیتابیس جدید"""
    text = update.message.text.strip()
    coupon_data = context.user_data.get('coupon_data', {})

    if 'code' not in coupon_data:
        # مرحله 1: کد کوپن
        if not re.match(r'^[A-Za-z0-9\-_]+$', text):
            await update.message.reply_text("❌ کد کوپن نامعتبر! فقط حروف، اعداد و خط تیره.")
            return

        # چک تکراری نبودن
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM coupons WHERE code = %s", (text,))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            await update.message.reply_text("❌ این کد قبلاً استفاده شده!")
            return
        cursor.close()
        conn.close()

        coupon_data['code'] = text
        context.user_data['coupon_data'] = coupon_data

        keyboard = [
            [InlineKeyboardButton("📊 درصد", callback_data="coupon_type_percent")],
            [InlineKeyboardButton("💰 مبلغ ثابت", callback_data="coupon_type_amount")],
            [InlineKeyboardButton("🔙 انصراف", callback_data="admin_coupons")]
        ]

        await update.message.reply_text(
            f"✅ کد: <code>{text}</code>\n\nنوع تخفیف را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    elif 'discount_value' not in coupon_data:
        # مرحله 2: مقدار تخفیف
        try:
            value = int(text.replace(',', ''))
            coupon_data['discount_value'] = value
            context.user_data['coupon_data'] = coupon_data

            await update.message.reply_text(
                "✅ تخفیف ثبت شد.\n\nحداکثر تعداد استفاده را وارد کنید:\nبرای نامحدود 0 وارد کنید."
            )
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید!")

    elif 'max_uses' not in coupon_data:
        # مرحله 3: حداکثر استفاده
        try:
            value = int(text)
            coupon_data['max_uses'] = value if value > 0 else None
            context.user_data['coupon_data'] = coupon_data

            await update.message.reply_text(
                "تاریخ انقضا را وارد کنید (روز):\nمثال: 30 (برای 30 روز)\nبرای بدون انقضا 0 وارد کنید."
            )
        except ValueError:
            await update.message.reply_text("❌ عدد نامعتبر!")

    elif 'expires_days' not in coupon_data:
        # مرحله 4: تاریخ انقضا
        try:
            days = int(text)
            expires_at = None
            if days > 0:
                expires_at = datetime.now() + timedelta(days=days)

            conn = db.get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                INSERT INTO coupons
                (code, type, value, usage_limit, used_count, expires_at, is_active, created_at)
                VALUES (%s, %s, %s, %s, 0, %s, 1, NOW())
            """, (
                coupon_data['code'],
                coupon_data.get('discount_type', 'amount'),  # 'percent' یا 'amount'
                coupon_data['discount_value'],
                coupon_data.get('max_uses'),
                expires_at
            ))

            conn.commit()
            cursor.close()
            conn.close()

            await update.message.reply_text(
                f"✅ کوپن با موفقیت ایجاد شد!\n"
                f"🎟 کد: <code>{coupon_data['code']}</code>\n"
                f"💰 نوع: {coupon_data.get('discount_type', 'مبلغ ثابت')}\n"
                f"📊 حداکثر استفاده: {coupon_data.get('max_uses') or '∞'}\n"
                f"⏰ انقضا: {format_date(expires_at) if expires_at else 'بدون محدودیت'}",
                parse_mode='HTML'
            )

            context.user_data.pop('state', None)
            context.user_data.pop('coupon_data', None)

        except Exception as e:
            await update.message.reply_text(f"❌ خطا در ایجاد کوپن: {e}")




async def start_search_coupon(query, context):
    """شروع جستجوی کوپن"""
    context.user_data['state'] = WAITING_COUPON_SEARCH
    
    text = "🔍 <b>جستجوی کوپن</b>\n\n"
    text += "کد کوپن یا بخشی از آن را وارد کنید:\n"
    text += "مثال: <code>SUMMER</code>"
    
    keyboard = [[InlineKeyboardButton("🔙 انصراف", callback_data="admin_coupons")]]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

async def process_coupon_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """پردازش جستجوی کوپن"""
    search_term = update.message.text.strip().upper()
    
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    # جستجو در کدهای کوپن
    cursor.execute("""
        SELECT * FROM coupons 
        WHERE UPPER(code) LIKE %s 
        ORDER BY created_at DESC
        LIMIT 20
    """, (f'%{search_term}%',))
    
    coupons = cursor.fetchall()
    cursor.close()
    conn.close()
    
    if not coupons:
        await update.message.reply_text(
            f"❌ هیچ کوپنی با عبارت '<code>{search_term}</code>' یافت نشد.",
            parse_mode='HTML'
        )
        context.user_data.pop('state', None)
        return
    
    text = f"🔍 <b>نتایج جستجو برای:</b> <code>{search_term}</code>\n\n"
    text += f"📊 {len(coupons)} کوپن یافت شد:\n\n"
    
    keyboard = []
    
    for coupon in coupons:
        # وضعیت
        if not coupon['is_active']:
            status = "❌ غیرفعال"
        elif coupon.get('expires_at') and coupon['expires_at'] < datetime.now():
            status = "⏰ منقضی"
        elif coupon.get('max_uses') and coupon.get('used_count', 0) >= coupon['max_uses']:
            status = "🚫 تمام شده"
        else:
            status = "✅ فعال"
        
        # تخفیف
        if coupon.get('discount_percent'):
            discount = f"{coupon['discount_percent']}%"
        else:
            discount = format_price(coupon.get('discount_amount', 0))
        
        # استفاده
        used = coupon.get('used_count', 0)
        max_uses = coupon.get('max_uses') or '∞'
        
        text += f"{status} <code>{coupon['code']}</code>\n"
        text += f"   💰 تخفیف: {discount}\n"
        text += f"   📊 استفاده: {used}/{max_uses}\n"
        
        if coupon.get('expires_at'):
            text += f"   ⏰ انقضا: {format_date(coupon['expires_at'])}\n"
        
        if coupon.get('min_purchase_amount'):
            text += f"   🛒 حداقل خرید: {format_price(coupon['min_purchase_amount'])}\n"
        
        text += "\n"
        
        # دکمه برای هر کوپن
        keyboard.append([
            InlineKeyboardButton(
                f"🎟 {coupon['code']} - جزئیات",
                callback_data=f"admin_coupon_detail_{coupon['id']}"
            )
        ])
    
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_coupons")])
    
    await update.message.reply_text(
        text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML'
    )
    
    context.user_data.pop('state', None)


async def show_coupon_full_stats(query, coupon_id):
    """آمار کامل یک کوپن با جدول جدید"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    # اطلاعات کوپن
    cursor.execute("SELECT * FROM coupons WHERE id = %s", (coupon_id,))
    coupon = cursor.fetchone()
    if not coupon:
        cursor.close()
        conn.close()
        await query.message.edit_text("❌ کوپن یافت نشد.")
        return

    # آمار استفاده روزانه (7 روز اخیر)
    cursor.execute("""
        SELECT DATE(used_at) as date, COUNT(*) as uses, COUNT(DISTINCT user_id) as unique_users
        FROM coupon_usage
        WHERE coupon_id = %s AND used_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)
        GROUP BY DATE(used_at)
        ORDER BY date DESC
    """, (coupon_id,))
    daily_usage = cursor.fetchall()

    # کاربران برتر
    cursor.execute("""
        SELECT u.first_name, u.username, COUNT(cu.id) as use_count
        FROM coupon_usage cu
        LEFT JOIN users u ON cu.user_id = u.user_id
        WHERE cu.coupon_id = %s
        GROUP BY cu.user_id
        ORDER BY use_count DESC
        LIMIT 10
    """, (coupon_id,))
    top_users = cursor.fetchall()

    cursor.close()
    conn.close()

    discount_text = f"{coupon['value']}%" if coupon['type'] == 'percent' else format_price(coupon['value'])

    text = f"📊 <b>آمار کامل کوپن</b>\n\n"
    text += f"🎟 کد: <code>{coupon['code']}</code>\n"
    text += f"💰 تخفیف: {discount_text}\n\n"

    if daily_usage:
        text += "<b>📅 روند 7 روز اخیر:</b>\n"
        for stat in daily_usage:
            jalali = gregorian_to_jalali(stat['date'])
            text += f"• {jalali}: {stat['uses']} استفاده ({stat['unique_users']} کاربر)\n"
        text += "\n"

    if top_users:
        text += "<b>👥 کاربران پرمصرف:</b>\n"
        for idx, user in enumerate(top_users[:5], 1):
            name = user.get('first_name', 'نامشخص')
            username = f"@{user['username']}" if user.get('username') else ''
            text += f"{idx}. {name} {username} | 📊 {user['use_count']} استفاده\n"
        text += "\n"

    keyboard = [
        [InlineKeyboardButton("📥 Excel این کوپن", callback_data=f"admin_export_coupon_{coupon_id}")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data=f"admin_coupon_detail_{coupon_id}")]
    ]

    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')




async def show_coupon_detail(query, coupon_id):
    """نمایش جزئیات کامل کوپن طبق دیتابیس جدید"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    # اطلاعات کوپن
    cursor.execute("SELECT * FROM coupons WHERE id = %s", (coupon_id,))
    coupon = cursor.fetchone()

    if not coupon:
        cursor.close()
        conn.close()
        await query.message.edit_text("❌ کوپن یافت نشد.")
        return

    # آخرین استفاده‌کنندگان
    cursor.execute("""
        SELECT cu.*, u.first_name, u.username
        FROM coupon_usage cu
        LEFT JOIN users u ON cu.user_id = u.user_id
        WHERE cu.coupon_id = %s
        ORDER BY cu.used_at DESC
        LIMIT 10
    """, (coupon_id,))
    usage_list = cursor.fetchall()

    # آمار کلی استفاده
    cursor.execute("""
        SELECT 
            COUNT(*) as total_uses,
            COUNT(DISTINCT user_id) as unique_users
        FROM coupon_usage
        WHERE coupon_id = %s
    """, (coupon_id,))
    stats = cursor.fetchone()

    cursor.close()
    conn.close()

    # وضعیت
    if not coupon['is_active']:
        status = "❌ غیرفعال"
        status_emoji = "❌"
    elif coupon.get('expires_at') and coupon['expires_at'] < datetime.now():
        status = "⏰ منقضی شده"
        status_emoji = "⏰"
    elif coupon.get('usage_limit') and coupon.get('used_count', 0) >= coupon.get('usage_limit'):
        status = "🚫 ظرفیت تکمیل"
        status_emoji = "🚫"
    else:
        status = "✅ فعال"
        status_emoji = "✅"

    discount_text = f"{coupon['value']}%" if coupon['type'] == 'percent' else format_price(coupon['value'])

    text = f"🎟 <b>جزئیات کوپن</b>\n\n"
    text += f"{status_emoji} <b>کد:</b> <code>{coupon['code']}</code>\n"
    text += f"📊 <b>وضعیت:</b> {status}\n"
    text += f"💰 <b>تخفیف:</b> {discount_text}\n"
    text += f"📈 <b>استفاده:</b> {coupon['used_count']}/{coupon['usage_limit'] or '∞'}\n"
    text += f"⏰ <b>انقضا:</b> {format_date(coupon['expires_at']) if coupon.get('expires_at') else 'بدون محدودیت'}\n"
    text += f"📅 <b>ایجاد:</b> {format_date(coupon['created_at'])}\n\n"

    text += f"<b>📊 آمار کلی استفاده:</b>\n"
    text += f"• کل استفاده: {stats['total_uses'] or 0}\n"
    text += f"• کاربران منحصر: {stats['unique_users'] or 0}\n\n"

    if usage_list:
        text += f"<b>👥 آخرین استفاده‌کنندگان:</b>\n"
        for usage in usage_list[:5]:
            name = usage.get('first_name', 'نامشخص')
            username = f"@{usage['username']}" if usage.get('username') else ''
            used_at = format_date(usage['used_at'])
            text += f"• {name} {username} | {used_at}\n"

    keyboard = []
    if coupon['is_active']:
        keyboard.append([
            InlineKeyboardButton("❌ غیرفعال کردن", callback_data=f"admin_coupon_disable_{coupon_id}"),
            InlineKeyboardButton("✏️ ویرایش", callback_data=f"admin_coupon_edit_{coupon_id}")
        ])
    else:
        keyboard.append([
            InlineKeyboardButton("✅ فعال کردن", callback_data=f"admin_coupon_enable_{coupon_id}"),
            InlineKeyboardButton("✏️ ویرایش", callback_data=f"admin_coupon_edit_{coupon_id}")
        ])
    keyboard.append([
        InlineKeyboardButton("📊 آمار کامل", callback_data=f"admin_coupon_fullstats_{coupon_id}"),
        InlineKeyboardButton("🗑 حذف", callback_data=f"admin_coupon_delete_{coupon_id}")
    ])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_coupons")])

    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_coupon_stats(query):
    """نمایش آمار کامل استفاده از کوپن‌ها طبق دیتابیس جدید"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    # آمار کلی
    cursor.execute("""
        SELECT 
            COUNT(*) as total_coupons,
            SUM(used_count) as total_uses,
            COUNT(DISTINCT cu.user_id) as unique_users
        FROM coupons c
        LEFT JOIN coupon_usage cu ON c.id = cu.coupon_id
    """)
    overall_stats = cursor.fetchone()

    # پرکاربردترین کوپن‌ها
    cursor.execute("""
        SELECT 
            c.code,
            c.type,
            c.value,
            COUNT(cu.id) as use_count,
            COUNT(DISTINCT cu.user_id) as unique_users
        FROM coupons c
        LEFT JOIN coupon_usage cu ON c.id = cu.coupon_id
        GROUP BY c.id
        ORDER BY use_count DESC
        LIMIT 10
    """)
    top_coupons = cursor.fetchall()

    # کوپن‌های نزدیک به اتمام ظرفیت
    cursor.execute("""
        SELECT code, used_count, usage_limit, (usage_limit - used_count) as remaining
        FROM coupons
        WHERE usage_limit IS NOT NULL 
        AND used_count >= (usage_limit * 0.8)
        AND is_active = 1
        ORDER BY remaining ASC
        LIMIT 5
    """)
    near_limit = cursor.fetchall()

    cursor.close()
    conn.close()

    # ساخت متن پیام
    text = "📊 <b>آمار کامل کوپن‌های تخفیف</b>\n\n"
    text += "<b>📈 آمار کلی:</b>\n"
    text += f"• تعداد کوپن‌ها: {overall_stats['total_coupons'] or 0}\n"
    text += f"• کل استفاده‌ها: {overall_stats['total_uses'] or 0}\n"
    text += f"• کاربران منحصر: {overall_stats['unique_users'] or 0}\n\n"

    # پرکاربردترین کوپن‌ها
    if top_coupons:
        text += "<b>🏆 پرکاربردترین کوپن‌ها:</b>\n"
        for idx, coupon in enumerate(top_coupons[:5], 1):
            medal = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉" if idx == 3 else f"{idx}."
            discount = f"{coupon['value']}%" if coupon['type'] == 'percent' else format_price(coupon['value'])
            text += f"{medal} <code>{coupon['code']}</code> ({discount}) | 📊 {coupon['use_count']} استفاده | 👥 {coupon['unique_users']} کاربر\n"
        text += "\n"

    # نزدیک به اتمام ظرفیت
    if near_limit:
        text += "<b>⚠️ نزدیک به اتمام ظرفیت:</b>\n"
        for c in near_limit:
            text += f"• <code>{c['code']}</code>: {c['used_count']}/{c['usage_limit']} (باقیمانده: {c['remaining']})\n"
        text += "\n"

    keyboard = [
        [InlineKeyboardButton("📥 دانلود Excel", callback_data="admin_export_coupon_stats")],
        [InlineKeyboardButton("📊 نمودار", callback_data="admin_coupon_chart")],
        [InlineKeyboardButton("🔄 بروزرسانی", callback_data="admin_coupon_stats")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_coupons")]
    ]

    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')



async def export_coupon_stats_excel(query, context, coupon_id=None):
    """خروجی Excel آمار کوپن‌ها"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        from io import BytesIO
        
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "آمار کوپن‌ها"
        
        # هدرها
        headers = ['ردیف', 'کد کوپن', 'تخفیف', 'تعداد استفاده', 'کاربران منحصر', 'کل تخفیف', 'وضعیت']
        ws.append(headers)
        
        # استایل هدر
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # داده‌ها
        if coupon_id:
            # یک کوپن خاص
            cursor.execute("""
                SELECT 
                    c.*,
                    COUNT(o.id) as use_count,
                    COUNT(DISTINCT o.user_id) as unique_users,
                    SUM(o.discount_amount) as total_discount
                FROM coupons c
                LEFT JOIN orders o ON c.code = o.coupon_code
                WHERE c.id = %s
                GROUP BY c.id
            """, (coupon_id,))
        else:
            # همه کوپن‌ها
            cursor.execute("""
                SELECT 
                    c.*,
                    COUNT(o.id) as use_count,
                    COUNT(DISTINCT o.user_id) as unique_users,
                    SUM(o.discount_amount) as total_discount
                FROM coupons c
                LEFT JOIN orders o ON c.code = o.coupon_code
                GROUP BY c.id
                ORDER BY use_count DESC
            """)
        
        coupons = cursor.fetchall()
        
        for idx, coupon in enumerate(coupons, 1):
            if coupon.get('discount_percent'):
                discount = f"{coupon['discount_percent']}%"
            else:
                discount = f"{coupon.get('discount_amount', 0):,} تومان"
            
            status = "فعال" if coupon['is_active'] else "غیرفعال"
            
            ws.append([
                idx,
                coupon['code'],
                discount,
                coupon['use_count'] or 0,
                coupon['unique_users'] or 0,
                coupon['total_discount'] or 0,
                status
            ])
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        cursor.close()
        conn.close()
        
        # ذخیره
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"coupon_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=excel_file,
            filename=filename,
            caption=f"📊 آمار کوپن‌ها\n📅 {datetime.now().strftime('%Y/%m/%d %H:%M')}"
        )
        
        await query.answer("✅ فایل ارسال شد", show_alert=False)
        
    except ImportError:
        await query.answer("❌ کتابخانه openpyxl نصب نیست", show_alert=True)
    except Exception as e:
        logger.error(f"خطا در export Excel: {e}")
        await query.answer("❌ خطا در ایجاد فایل", show_alert=True)






async def show_admin_coupons_list(query, page=1, filter_type='all'):
    """نمایش لیست کوپن‌ها با جدول جدید"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    per_page = 10
    offset = (page - 1) * per_page

    where_clause = ""
    if filter_type == 'active':
        where_clause = "WHERE is_active = 1 AND (expires_at IS NULL OR expires_at > NOW())"
    elif filter_type == 'expired':
        where_clause = "WHERE expires_at < NOW()"
    elif filter_type == 'disabled':
        where_clause = "WHERE is_active = 0"

    cursor.execute(f"""
        SELECT * FROM coupons
        {where_clause}
        ORDER BY created_at DESC
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    coupons = cursor.fetchall()

    cursor.execute(f"SELECT COUNT(*) as total FROM coupons {where_clause}")
    total = cursor.fetchone()['total']

    cursor.close()
    conn.close()

    total_pages = (total + per_page - 1) // per_page
    filter_names = {'all': 'همه', 'active': 'فعال', 'expired': 'منقضی', 'disabled': 'غیرفعال'}

    text = f"🎟 <b>لیست کوپن‌ها ({filter_names[filter_type]})</b>\n\n"
    text += f"📄 صفحه {page} از {total_pages}\n\n"

    if coupons:
        for coupon in coupons:
            status = "✅" if coupon['is_active'] else "❌"
            if coupon.get('expires_at') and coupon['expires_at'] < datetime.now():
                status = "⏰"

            discount = f"{coupon['value']}%" if coupon['type'] == 'percent' else format_price(coupon['value'])
            text += f"{status} <code>{coupon['code']}</code> | 💰 {discount} | 📊 {coupon['used_count']}/{coupon['usage_limit'] or '∞'}\n"
    else:
        text += "❌ کوپنی یافت نشد."

    keyboard = [
        [
            InlineKeyboardButton("همه", callback_data="admin_coupons_filter_all"),
            InlineKeyboardButton("✅ فعال", callback_data="admin_coupons_filter_active"),
            InlineKeyboardButton("⏰ منقضی", callback_data="admin_coupons_filter_expired"),
            InlineKeyboardButton("❌ غیرفعال", callback_data="admin_coupons_filter_disabled")
        ]
    ]
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("◀️", callback_data=f"admin_coupons_list_{filter_type}_{page-1}"))
    nav_buttons.append(InlineKeyboardButton(f"{page}/{total_pages}", callback_data="noop"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("▶️", callback_data=f"admin_coupons_list_{filter_type}_{page+1}"))
    keyboard.append(nav_buttons)
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_coupons")])

    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def show_admin_campaigns_menu(query):
    """منوی مدیریت کمپین‌های تبلیغاتی"""
    text = "🎉 <b>مدیریت کمپین‌های تبلیغاتی</b>\n\n"
    text += "⚠️ این بخش در حال توسعه است.\n\n"
    text += "کمپین‌های آینده:\n"
    text += "• 🎁 جشنواره تخفیف\n"
    text += "• 🎯 تبلیغات هدفمند\n"
    text += "• 📧 ایمیل مارکتینگ\n"
    text += "• 📱 نوتیفیکیشن هوشمند"
    
    keyboard = [
        [InlineKeyboardButton("🔜 به زودی", callback_data="noop")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_financial")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')



async def toggle_coupon_status(query, coupon_id, enable=True):
    """فعال یا غیرفعال کردن کوپن"""
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute(
        "UPDATE coupons SET is_active = %s WHERE id = %s",
        (1 if enable else 0, coupon_id)
    )
    
    conn.commit()
    
    # دریافت کد کوپن
    cursor.execute("SELECT code FROM coupons WHERE id = %s", (coupon_id,))
    result = cursor.fetchone()
    coupon_code = result[0] if result else "نامشخص"
    
    cursor.close()
    conn.close()
    
    status_text = "فعال" if enable else "غیرفعال"
    
    await query.answer(f"✅ کوپن {coupon_code} {status_text} شد", show_alert=True)
    
    log_admin_action(
        query.from_user.id,
        'toggle_coupon',
        None,
        f"{status_text} کردن کوپن {coupon_code}"
    )
    
    # بازگشت به جزئیات
    await show_coupon_detail(query, coupon_id)

async def delete_coupon(query, coupon_id):
    """حذف کوپن"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    # دریافت اطلاعات کوپن
    cursor.execute("SELECT code FROM coupons WHERE id = %s", (coupon_id,))
    coupon = cursor.fetchone()
    
    if not coupon:
        cursor.close()
        conn.close()
        await query.answer("❌ کوپن یافت نشد", show_alert=True)
        return
    
    # حذف
    cursor.execute("DELETE FROM coupons WHERE id = %s", (coupon_id,))
    conn.commit()
    cursor.close()
    conn.close()
    
    await query.answer(f"✅ کوپن {coupon['code']} حذف شد", show_alert=True)
    
    log_admin_action(
        query.from_user.id,
        'delete_coupon',
        None,
        f"حذف کوپن {coupon['code']}"
    )
    
    # بازگشت به لیست
    await show_admin_coupons_menu(query)

async def toggle_coupon_status(query, coupon_id, enable=True):
    """فعال یا غیرفعال کردن کوپن"""
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute(
        "UPDATE coupons SET is_active = %s WHERE id = %s",
        (1 if enable else 0, coupon_id)
    )
    
    conn.commit()
    
    # دریافت کد کوپن
    cursor.execute("SELECT code FROM coupons WHERE id = %s", (coupon_id,))
    result = cursor.fetchone()
    coupon_code = result[0] if result else "نامشخص"
    
    cursor.close()
    conn.close()
    
    status_text = "فعال" if enable else "غیرفعال"
    
    await query.answer(f"✅ کوپن {coupon_code} {status_text} شد", show_alert=True)
    
    log_admin_action(
        query.from_user.id,
        'toggle_coupon',
        None,
        f"{status_text} کردن کوپن {coupon_code}"
    )
    
    # بازگشت به جزئیات
    await show_coupon_detail(query, coupon_id)

async def delete_coupon(query, coupon_id):
    """حذف کوپن"""
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    
    # دریافت اطلاعات کوپن
    cursor.execute("SELECT code FROM coupons WHERE id = %s", (coupon_id,))
    coupon = cursor.fetchone()
    
    if not coupon:
        cursor.close()
        conn.close()
        await query.answer("❌ کوپن یافت نشد", show_alert=True)
        return
    
    # حذف
    cursor.execute("DELETE FROM coupons WHERE id = %s", (coupon_id,))
    conn.commit()
    cursor.close()
    conn.close()
    
    await query.answer(f"✅ کوپن {coupon['code']} حذف شد", show_alert=True)
    
    log_admin_action(
        query.from_user.id,
        'delete_coupon',
        None,
        f"حذف کوپن {coupon['code']}"
    )
    
    # بازگشت به لیست
    await show_admin_coupons_menu(query)



async def export_financial_excel(query, context, export_type='all'):
    """خروجی Excel گزارشات مالی"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        
        # ایجاد Workbook
        wb = openpyxl.Workbook()
        
        if export_type in ['all', 'transactions']:
            ws_trans = wb.active
            ws_trans.title = "تراکنش‌ها"
            
            # هدرها
            headers = ['ردیف', 'کاربر', 'نوع', 'مبلغ', 'توضیحات', 'تاریخ']
            ws_trans.append(headers)
            
            # استایل هدر
            for cell in ws_trans[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # داده‌ها
            cursor.execute("""
                SELECT t.*, u.first_name, u.username 
                FROM transactions t
                LEFT JOIN users u ON t.user_id = u.user_id
                ORDER BY t.created_at DESC
                LIMIT 1000
            """)
            transactions = cursor.fetchall()
            
            for idx, trans in enumerate(transactions, 1):
                ws_trans.append([
                    idx,
                    f"{trans.get('first_name', 'نامشخص')} (@{trans.get('username', 'بدون یوزر')})",
                    trans['type'],
                    trans['amount'],
                    trans.get('description', ''),
                    trans['created_at'].strftime('%Y-%m-%d %H:%M')
                ])
        
        if export_type in ['all', 'payments']:
            ws_pay = wb.create_sheet("پرداخت‌ها")
            
            headers = ['ردیف', 'کاربر', 'مبلغ', 'وضعیت', 'Authority', 'تاریخ']
            ws_pay.append(headers)
            
            for cell in ws_pay[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            cursor.execute("""
                SELECT p.*, u.first_name, u.username
                FROM payments p
                LEFT JOIN users u ON p.user_id = u.user_id
                ORDER BY p.created_at DESC
                LIMIT 1000
            """)
            payments = cursor.fetchall()
            
            for idx, pay in enumerate(payments, 1):
                ws_pay.append([
                    idx,
                    f"{pay.get('first_name', 'نامشخص')}",
                    pay['amount'],
                    pay['status'],
                    pay.get('authority', 'N/A'),
                    pay['created_at'].strftime('%Y-%m-%d %H:%M')
                ])
        
        cursor.close()
        conn.close()
        
        # ذخیره در BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # ارسال فایل
        filename = f"financial_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=excel_file,
            filename=filename,
            caption=f"📥 گزارش مالی\n📅 {datetime.now().strftime('%Y/%m/%d %H:%M')}"
        )
        
        await query.answer("✅ فایل ارسال شد", show_alert=False)
        
    except ImportError:
        await query.answer("❌ کتابخانه openpyxl نصب نیست", show_alert=True)
    except Exception as e:
        logger.error(f"خطا در export Excel: {e}")
        await query.answer("❌ خطا در ایجاد فایل", show_alert=True)




async def show_admin_referral_menu(query):
    text = "🎁 <b>مدیریت رفرال</b>\n\nگزینه مورد نظر را انتخاب کنید:"
    
    inviter_reward = get_setting('referral_inviter_reward', '10000')
    invited_reward = get_setting('referral_invited_reward', '5000')
    
    text += f"\n💰 پاداش دعوت‌کننده: {format_price(int(inviter_reward))}\n"
    text += f"💰 پاداش دعوت‌شده: {format_price(int(invited_reward))}\n"
    
    keyboard = [
        [InlineKeyboardButton("🏆 برترین دعوت‌کنندگان", callback_data="admin_top_referrers")],
        [InlineKeyboardButton("⚙️ تنظیم پاداش دعوت‌کننده", callback_data="admin_set_inviter_reward")],
        [InlineKeyboardButton("⚙️ تنظیم پاداش دعوت‌شده", callback_data="admin_set_invited_reward")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_panel")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

async def show_admin_settings_menu(query):
    text = "⚙️ <b>تنظیمات سیستم</b>\n\nگزینه مورد نظر را انتخاب کنید:"
    
    keyboard = [
        [InlineKeyboardButton("🔧 تنظیمات Marzban", callback_data="admin_marzban_settings")],
        [InlineKeyboardButton("💳 تنظیمات زرین‌پال", callback_data="admin_zarinpal_settings")],
        [InlineKeyboardButton("📝 ویرایش پیام خوش‌آمد", callback_data="admin_welcome_message")],
        [InlineKeyboardButton("🔌 تست اتصال Marzban", callback_data="admin_test_marzban")],
        [InlineKeyboardButton("📜 مشاهده لاگ‌ها", callback_data="admin_logs")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data="admin_panel")]
    ]
    
    await safe_edit_message(query, text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')

# ==================== MESSAGE HANDLERS ====================

async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state = context.user_data.get('state')
    
    if state == WAITING_WALLET_CHARGE_AMOUNT:
        try:
            amount = int(update.message.text.replace(',', ''))
            if amount < 10000:
                await update.message.reply_text("❌ حداقل مبلغ شارژ 10,000 تومان است.")
                return
            
            # TODO: پیاده‌سازی درگاه پرداخت
            await update.message.reply_text(f"✅ درخواست شارژ {format_price(amount)} ثبت شد.\n\n⚠️ این قابلیت در حال توسعه است.")
            context.user_data.pop('state', None)
            
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید.")
    
    elif state == WAITING_BROADCAST_MESSAGE:
        if user_id not in ADMIN_IDS:
            return
        
        message_text = update.message.text
        all_users = get_all_users()
        
        await update.message.reply_text(f"📤 در حال ارسال پیام به {len(all_users)} کاربر...")
        
        success = 0
        failed = 0
        
        for user in all_users:
            try:
                await context.bot.send_message(chat_id=user['user_id'], text=message_text)
                success += 1
                await asyncio.sleep(0.05)  # جلوگیری از rate limit
            except:
                failed += 1
        
        await update.message.reply_text(f"✅ ارسال پیام تکمیل شد.\n\n✅ موفق: {success}\n❌ ناموفق: {failed}")
        log_admin_action(user_id, 'broadcast', None, f"ارسال به {success} کاربر")
        context.user_data.pop('state', None)
    elif state == WAITING_BULK_BALANCE_AMOUNT:
        try:
            amount = int(update.message.text.replace(',', ''))
            
            if amount < 1000:
                await update.message.reply_text("❌ حداقل مبلغ 1,000 تومان است.")
                return
            
            tag = context.user_data.get('bulk_balance_tag')
            user_count = get_users_count(tag=tag if tag != 'all' else None)
            total_amount = amount * user_count
            
            context.user_data['bulk_balance_amount'] = amount
            context.user_data['state'] = WAITING_BULK_BALANCE_REASON
            
            await update.message.reply_text(
                f"💰 <b>تایید افزایش موجودی گروهی</b>\n\n"
                f"🏷 تگ: <code>{tag}</code>\n"
                f"👥 تعداد کاربران: {user_count}\n"
                f"💵 مبلغ هر کاربر: {format_price(amount)}\n"
                f"💰 مجموع کل: {format_price(total_amount)}\n\n"
                f"📝 دلیل این افزایش را وارد کنید:",
                parse_mode='HTML'
            )
        
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید.")
    
    # ==================== BULK BALANCE REASON ====================
    elif state == WAITING_BULK_BALANCE_REASON:
        reason = update.message.text.strip()
        
        if len(reason) < 3:
            await update.message.reply_text("❌ دلیل باید حداقل 3 کاراکتر باشد.")
            return
        
        tag = context.user_data.get('bulk_balance_tag')
        amount = context.user_data.get('bulk_balance_amount')
        user_count = get_users_count(tag=tag if tag != 'all' else None)
        
        keyboard = [
            [InlineKeyboardButton("✅ تایید و اجرا", callback_data="admin_confirm_bulk_balance")],
            [InlineKeyboardButton("❌ انصراف", callback_data="admin_users")]
        ]
        
        context.user_data['bulk_balance_reason'] = reason
        
        await update.message.reply_text(
            f"⚠️ <b>تایید نهایی</b>\n\n"
            f"🏷 تگ: <code>{tag}</code>\n"
            f"👥 کاربران: {user_count}\n"
            f"💵 مبلغ: {format_price(amount)}\n"
            f"💰 جمع: {format_price(amount * user_count)}\n"
            f"📝 دلیل: {reason}\n\n"
            f"⚠️ این عمل قابل بازگشت نیست!",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )
    
    # ==================== USER TAG (CREATE/RENAME) ====================
    elif state == WAITING_USER_TAG:
        tag_name = update.message.text.strip().lower()
        
        # اعتبارسنجی
        if not tag_name.isalnum():
            await update.message.reply_text("❌ نام تگ فقط باید شامل حروف و اعداد انگلیسی باشد.")
            return
        
        if len(tag_name) < 2 or len(tag_name) > 20:
            await update.message.reply_text("❌ طول نام تگ باید بین 2 تا 20 کاراکتر باشد.")
            return
        
        # بررسی ایجاد تگ جدید یا تغییر نام
        if context.user_data.get('creating_new_tag'):
            # ذخیره تگ جدید (فعلاً فقط اعلام موفقیت)
            await update.message.reply_text(
                f"✅ تگ <code>{tag_name}</code> ایجاد شد!\n\n"
                f"حالا می‌توانید از بخش مدیریت کاربران، این تگ را به کاربران اختصاص دهید.",
                parse_mode='HTML'
            )
            context.user_data.pop('creating_new_tag', None)
        
        elif context.user_data.get('renaming_tag'):
            old_tag = context.user_data.get('renaming_tag')
            
            # تغییر نام تگ
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE users SET user_tag = %s WHERE user_tag = %s",
                (tag_name, old_tag)
            )
            affected = cursor.rowcount
            conn.commit()
            cursor.close()
            conn.close()
            
            log_admin_action(user_id, 'rename_tag', 0, f"تغییر نام {old_tag} به {tag_name}")
            
            await update.message.reply_text(
                f"✅ نام تگ تغییر یافت!\n\n"
                f"🏷 از <code>{old_tag}</code> به <code>{tag_name}</code>\n"
                f"👥 {affected} کاربر بروزرسانی شد.",
                parse_mode='HTML'
            )
            context.user_data.pop('renaming_tag', None)
        
        context.user_data.pop('state', None)

    elif state == WAITING_USER_SEARCH:
        if user_id not in ADMIN_IDS:
            return
        
        search = update.message.text.strip()
        users = get_all_users(limit=10, search=search)
        
        if not users:
            await update.message.reply_text("❌ کاربری یافت نشد.")
            context.user_data.pop('state', None)
            return
        
        text = f"🔍 <b>نتایج جستجو برای «{search}»:</b>\n\n"
        keyboard = []
        
        for u in users:
            text += f"👤 {u['first_name']} (@{u['username'] or 'بدون نام کاربری'})\n"
            text += f"   🆔 {u['user_id']}\n\n"
            
            keyboard.append([InlineKeyboardButton(
                f"👤 {u['first_name']} - {u['user_id']}",
                callback_data=f"admin_view_user_{u['user_id']}"
            )])
        
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data="admin_users")])
        
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
        context.user_data.pop('state', None)
        # ==================== WAITING BALANCE AMOUNT (ADMIN) ====================
    elif state == WAITING_BALANCE_AMOUNT:
        if user_id not in ADMIN_IDS:
            return
    
        try:
            amount = int(update.message.text.replace(',', '').replace('+', ''))
            target_user_id = context.user_data.get('target_user_id')
        
            if not target_user_id:
                await update.message.reply_text("❌ خطا: کاربر هدف مشخص نشده.")
                context.user_data.pop('state', None)
                return
        
            target_user = get_user(target_user_id)
        
            if not target_user:
                await update.message.reply_text("❌ کاربر یافت نشد.")
                context.user_data.pop('state', None)
                return
        
        # محاسبه موجودی جدید
            new_balance = target_user['balance'] + amount
        
            if new_balance < 0:
                await update.message.reply_text(
                    f"❌ خطا: موجودی نمی‌تواند منفی شود!\n\n"
                    f"💵 موجودی فعلی: {format_price(target_user['balance'])}\n"
                    f"🔻 کاهش درخواستی: {format_price(abs(amount))}\n"
                    f"⚠️ حداکثر می‌توانید {format_price(target_user['balance'])} کاهش دهید."
                )
                return
        
            context.user_data['balance_amount'] = amount
            context.user_data['state'] = WAITING_BALANCE_REASON
        
            sign = "+" if amount >= 0 else ""
            action = "افزایش" if amount >= 0 else "کاهش"
        
            text = f"💰 <b>تایید تغییر موجودی</b>\n\n"
            text += f"👤 کاربر: {target_user['first_name']}\n"
            text += f"💵 موجودی فعلی: {format_price(target_user['balance'])}\n"
            text += f"📊 {action}: {sign}{format_price(amount)}\n"
            text += f"💰 موجودی جدید: {format_price(new_balance)}\n\n"
            text += "📝 دلیل این تغییر را وارد کنید:"
        
            await update.message.reply_text(text, parse_mode='HTML')
        
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید.")

# ==================== WAITING BALANCE REASON (ADMIN) ====================
    elif state == WAITING_BALANCE_REASON:
        if user_id not in ADMIN_IDS:
            return
    
        reason = update.message.text.strip()
    
        if len(reason) < 3:
            await update.message.reply_text("❌ دلیل باید حداقل 3 کاراکتر باشد.")
            return
    
        target_user_id = context.user_data.get('target_user_id')
        amount = context.user_data.get('balance_amount')
    
        target_user = get_user(target_user_id)
    
    # اعمال تغییرات
        update_user_balance(target_user_id, amount, reason, admin_id=user_id)
    
    # ثبت لاگ
        log_admin_action(
            user_id,
            'edit_balance',
            target_user_id,
            f"تغییر موجودی: {'+' if amount >= 0 else ''}{format_price(amount)} - {reason}"
        )
    
        sign = "+" if amount >= 0 else ""
    
        await update.message.reply_text(
            f"✅ <b>موجودی تغییر یافت!</b>\n\n"
            f"👤 کاربر: {target_user['first_name']}\n"
            f"📊 تغییر: {sign}{format_price(amount)}\n"
            f"💰 موجودی جدید: {format_price(target_user['balance'] + amount)}\n"
            f"📝 دلیل: {reason}",
            parse_mode='HTML'
        )
    
    # اطلاعیه به کاربر
        try:
            emoji = "🎉" if amount > 0 else "⚠️"
            action = "افزایش یافت" if amount > 0 else "کاهش یافت"
        
            await context.bot.send_message(
                target_user_id,
                f"{emoji} <b>تغییر موجودی</b>\n\n"
                f"موجودی حساب شما {sign}{format_price(amount)} {action}.\n\n"
                f"💰 موجودی جدید: {format_price(target_user['balance'] + amount)}\n"
                f"📝 دلیل: {reason}",
                parse_mode='HTML'
            )
        except Exception as e:
            logger.error(f"خطا در ارسال اطلاعیه به کاربر {target_user_id}: {e}")
    
    # پاکسازی
        context.user_data.pop('state', None)
        context.user_data.pop('target_user_id', None)
        context.user_data.pop('balance_amount', None)
        


    elif state == WAITING_REFERRAL_REWARD_INVITER:
        if user_id not in ADMIN_IDS:
            return
    
        try:
            amount = int(update.message.text.replace(',', ''))
        
            if amount < 0:
                await update.message.reply_text("❌ مبلغ نمی‌تواند منفی باشد.")
                return
        
            set_setting('referral_inviter_reward', str(amount))
        
            log_admin_action(user_id, 'change_setting', 0, f"پاداش دعوت‌کننده → {format_price(amount)}")
        
            await update.message.reply_text(
                f"✅ پاداش دعوت‌کننده تغییر یافت!\n\n"
                f"💰 مبلغ جدید: {format_price(amount)}"
            )
        
            context.user_data.pop('state', None)
        
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید.")

# ==================== WAITING REFERRAL REWARD INVITED ====================
    elif state == WAITING_REFERRAL_REWARD_INVITED:
        if user_id not in ADMIN_IDS:
            return
    
        try:
            amount = int(update.message.text.replace(',', ''))
        
            if amount < 0:
                await update.message.reply_text("❌ مبلغ نمی‌تواند منفی باشد.")
                return
        
            set_setting('referral_invited_reward', str(amount))
        
            log_admin_action(user_id, 'change_setting', 0, f"پاداش دعوت‌شده → {format_price(amount)}")
        
            await update.message.reply_text(
                f"✅ پاداش دعوت‌شده تغییر یافت!\n\n"
                f"💰 مبلغ جدید: {format_price(amount)}"
            )
        
            context.user_data.pop('state', None)
        
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید.")

    # در message_handler:
    elif state == WAITING_COUPON_DATA:
        await process_coupon_creation(update, context)


    elif state == WAITING_COUPON_SEARCH:
        await process_coupon_search(update, context)

    # ==================== WAITING MARZBAN URL ====================
    elif state == WAITING_MARZBAN_URL:
        if user_id not in ADMIN_IDS:
            return
    
        url = update.message.text.strip()
    
    # اعتبارسنجی URL
        if not url.startswith(('http://', 'https://')):
            await update.message.reply_text("❌ URL باید با http:// یا https:// شروع شود.")
            return
    
    # حذف slash انتهایی
        url = url.rstrip('/')
    
    # ذخیره
        set_setting('marzban_url', url)
    
    # بروزرسانی global marzban
        global marzban
        marzban = MarzbanAPI(
            url=url,
            username=get_setting('marzban_username', MARZBAN_USERNAME),
            password=get_setting('marzban_password', MARZBAN_PASSWORD)
        )
    
        log_admin_action(user_id, 'change_marzban_url', None, f"URL جدید: {url}")
    
        await update.message.reply_text(
            f"✅ URL مرزبان تغییر یافت!\n\n"
            f"🌐 URL جدید: <code>{url}</code>\n\n"
            f"💡 توصیه: حالا اتصال را تست کنید.",
            parse_mode='HTML'
            )   
    
        context.user_data.pop('state', None)

# ==================== WAITING MARZBAN USER ====================
    elif state == WAITING_MARZBAN_USER:
        
        if user_id not in ADMIN_IDS:
            return
    
        
        username = update.message.text.strip()
    
        if len(username) < 3:
            await update.message.reply_text("❌ نام کاربری باید حداقل 3 کاراکتر باشد.")
            return
    
        set_setting('marzban_username', username)
    
    # بروزرسانی global marzban
        
        marzban = MarzbanAPI(
            url=get_setting('marzban_url', MARZBAN_URL),
            username=username,
            password=get_setting('marzban_password', MARZBAN_PASSWORD)
        )
    
        log_admin_action(user_id, 'change_marzban_user', None, f"کاربر جدید: {username}")
    
        await update.message.reply_text(
            f"✅ نام کاربری مرزبان تغییر یافت!\n\n"
            f"👤 کاربر جدید: <code>{username}</code>",
            parse_mode='HTML'
        )
    
        context.user_data.pop('state', None)

# ==================== WAITING MARZBAN PASS ====================
    elif state == WAITING_MARZBAN_PASS:
        if user_id not in ADMIN_IDS:    
            return
    
        password = update.message.text.strip()
    
        if len(password) < 4:
            await update.message.reply_text("❌ رمز عبور باید حداقل 4 کاراکتر باشد.")
            return
    
        set_setting('marzban_password', password)
    
    # بروزرسانی global marzban
        
        marzban = MarzbanAPI(
            url=get_setting('marzban_url', MARZBAN_URL),
            username=get_setting('marzban_username', MARZBAN_USERNAME),
            password=password
        )
    
    # حذف پیام حاوی رمز عبور
        try:
            await update.message.delete()
        except:
            pass
    
        log_admin_action(user_id, 'change_marzban_pass', None, "رمز عبور تغییر یافت")
    
        await context.bot.send_message(
            chat_id=user_id,
            text="✅ رمز عبور مرزبان تغییر یافت!\n\n🔒 رمز عبور به صورت امن ذخیره شد."
        )
    
        context.user_data.pop('state', None)

# ==================== WAITING MERCHANT ID ====================
    elif state == WAITING_MERCHANT_ID:
        if user_id not in ADMIN_IDS:
            return
    
        merchant_id = update.message.text.strip()
    
    # اعتبارسنجی فرمت UUID
        if len(merchant_id) != 36 or merchant_id.count('-') != 4:
            await update.message.reply_text(
                "❌ فرمت Merchant ID نامعتبر!\n\n"
                "فرمت صحیح: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"
            )
            return
    
        set_setting('zarinpal_merchant', merchant_id)
    
        log_admin_action(user_id, 'change_zarinpal_merchant', None, "Merchant ID تغییر یافت")
    
        await update.message.reply_text(
            f"✅ Merchant ID زرین‌پال تغییر یافت!\n\n"
            f"🔑 ID جدید: <code>{merchant_id}</code>\n\n"
            f"💡 توصیه: درگاه را تست کنید.",
            parse_mode='HTML'
        )
    
        context.user_data.pop('state', None)

# ==================== WAITING WELCOME MESSAGE ====================
    # ==================== WAITING WELCOME MESSAGE ====================
    elif state == WAITING_WELCOME_MESSAGE:
        if user_id not in ADMIN_IDS:
            return

        new_message = update.message.text.strip()

        if len(new_message) < 10:
            await update.message.reply_text("❌ پیام خوش‌آمدگویی باید حداقل 10 کاراکتر باشد.")
            return

        set_setting('welcome_message', new_message)
        log_admin_action(user_id, 'change_welcome_message', None, "پیام خوش‌آمد تغییر یافت")

    # پیش‌نمایش امن
    # import re

        def safe_preview(text, sample_user, sample_db_user, sample_referrer=None, invited_reward=0, inviter_reward=0):
            replacements = {
                'user_name': sample_user.get('first_name', 'کاربر'),
                'first_name': sample_user.get('first_name', 'کاربر'),
                'user_id': str(sample_user.get('id', 0)),
                'balance': format_price(sample_db_user.get('balance', 50000)),
                'referrer_name': sample_referrer.get('first_name', 'یک دوست') if sample_referrer else '',
                'referrer_username': f"@{sample_referrer['username']}" if sample_referrer and sample_referrer.get('username') else '',
                'invited_reward': format_price(invited_reward) if sample_referrer else '',
                'inviter_reward': format_price(inviter_reward) if sample_referrer else ''
            }

            def replace_var(match):
                var_name = match.group(1)
                return str(replacements.get(var_name, match.group(0)))

            result = re.sub(r'\{(\w+)\}', replace_var, text)

        # حذف بخش هدیه در صورت عدم وجود رفرال
            if not sample_referrer:
                result = re.sub(r'🎁 هدیه .*? از .*?\n?', '', result)

            return result

    # نمونه پیش‌نمایش با و بدون رفرال
        sample_user = {"first_name": "محمد", "id": 123456789}
        sample_db_user = {"balance": 50000}
        sample_referrer = {"first_name": "علی", "username": "ali123"}
        sample_invited_reward = 5000
        sample_inviter_reward = 10000

        preview_with_referral = safe_preview(
            new_message, sample_user, sample_db_user,
            sample_referrer, sample_invited_reward, sample_inviter_reward
        )
        preview_without_referral = safe_preview(
            new_message, sample_user, sample_db_user
        )

        await update.message.reply_text(
            f"✅ پیام خوش‌آمدگویی تغییر یافت!\n\n"
            f"<b>📌 پیش‌نمایش با لینک دعوت:</b>\n{preview_with_referral}\n\n"
            f"<b>📌 پیش‌نمایش بدون لینک دعوت:</b>\n{preview_without_referral}",
            parse_mode='HTML'
        )

        context.user_data.pop('state', None)
    

    if state == WAITING_WALLET_CHARGE_AMOUNT:
        try:
            amount = int(update.message.text.replace(',', ''))
            if amount < 10000:
                await update.message.reply_text("❌ حداقل مبلغ شارژ 10,000 تومان است.")
                return
        
        # ذخیره مبلغ در context
            context.user_data['charge_amount'] = amount
        
            text = f"💰 <b>شارژ کیف پول</b>\n\n"
            text += f"💵 مبلغ: {format_price(amount)}\n\n"
            text += "روش پرداخت را انتخاب کنید:"
        
            keyboard = [
                [InlineKeyboardButton("💳 پرداخت آنلاین", callback_data="charge_online")],
                [InlineKeyboardButton("🔙 انصراف", callback_data="wallet")]
            ]
        
            await update.message.reply_text(
                text,
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode='HTML'
            )
        
            context.user_data.pop('state', None)
        
        except ValueError:
            await update.message.reply_text("❌ لطفاً یک عدد معتبر وارد کنید.")



# ==================== ZARINPAL CALLBACK WEBSERVER ====================

async def zarinpal_callback(request):
    """✅ دریافت callback از زرین‌پال + لاگ کامل"""
    
    # ✅ لاگ تمام پارامترهای دریافتی
    logger.info("=" * 50)
    logger.info("📥 ZARINPAL CALLBACK RECEIVED!")
    logger.info(f"Full URL: {request.url}")
    logger.info(f"Query params: {dict(request.query)}")
    logger.info("=" * 50)
    
    authority = request.query.get('Authority')
    status = request.query.get('Status')

    logger.info(f"Authority: {authority}")
    logger.info(f"Status: {status}")

    # مسیر صفحات HTML
    import os
    current_dir = os.path.dirname(os.path.abspath(__file__))
    success_path = os.path.join(current_dir, 'templates', 'success.html')
    failure_path = os.path.join(current_dir, 'templates', 'failure.html')

    if status == 'OK' and authority:
        logger.info(f"✅ Payment successful! Processing authority: {authority}")
        
        # ✅ پردازش async
        asyncio.create_task(process_successful_payment(authority))
        
        # 📄 نمایش صفحه موفق
        try:
            with open(success_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            logger.info("✅ Success page loaded and sent")
            return web.Response(text=html_content, content_type='text/html', charset='utf-8')
        except FileNotFoundError:
            logger.error(f"❌ Success page not found: {success_path}")
            html = """
            <!DOCTYPE html>
            <html lang="fa" dir="rtl">
            <head><meta charset="UTF-8"><title>پرداخت موفق</title></head>
            <body style="font-family:Tahoma;text-align:center;padding:50px;">
                <h1 style="color:green;">✅ پرداخت موفق!</h1>
                <p>به ربات تلگرام مراجعه کنید</p>
            </body>
            </html>
            """
            return web.Response(text=html, content_type='text/html', charset='utf-8')
    
    else:
        logger.warning(f"❌ Payment failed or cancelled. Authority: {authority}, Status: {status}")
        
        if authority:
            update_payment_status(authority, 'failed')
        
        # 📄 نمایش صفحه ناموفق
        try:
            with open(failure_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            logger.info("✅ Failure page loaded and sent")
            return web.Response(text=html_content, content_type='text/html', charset='utf-8')
        except FileNotFoundError:
            logger.error(f"❌ Failure page not found: {failure_path}")
            html = """
            <!DOCTYPE html>
            <html lang="fa" dir="rtl">
            <head><meta charset="UTF-8"><title>پرداخت ناموفق</title></head>
            <body style="font-family:Tahoma;text-align:center;padding:50px;">
                <h1 style="color:red;">❌ پرداخت ناموفق</h1>
                <p>لطفاً مجدداً تلاش کنید</p>
            </body>
            </html>
            """
            return web.Response(text=html, content_type='text/html', charset='utf-8')



async def process_successful_payment(authority: str):
    """پردازش پرداخت موفق"""
    try:
        payment = get_payment_by_authority(authority)
        if not payment:
            logger.error(f"Payment not found: {authority}")
            return

        if payment['status'] == 'success':
            logger.info(f"Payment already processed: {authority}")
            return

        # تایید با زرین‌پال
        merchant_id = get_setting('zarinpal_merchant', ZARINPAL_MERCHANT)
        zp = ZarinPal(merchant_id, ZARINPAL_SANDBOX)
        
        verify_result = zp.verify_payment(authority, payment['amount'])

        if verify_result.get('data', {}).get('code') == 100:
            ref_id = verify_result['data']['ref_id']
            update_payment_status(authority, 'success', ref_id)

            user_id = payment['user_id']

            if payment['payment_type'] == 'package':
                # ساخت سرویس
                pkg = PACKAGES.get(payment['package_id'])
                if pkg:
                    # لاجیک ساخت سرویس...
                    pass

            elif payment['payment_type'] == 'wallet':
                # شارژ کیف پول
                update_user_balance(user_id, payment['amount'] // 10, f"شارژ - کد: {ref_id}")

            logger.info(f"✅ Payment processed: {authority}")

    except Exception as e:
        logger.error(f"Error processing payment: {e}")





async def send_service_activation_message(user_id, payment, ref_id):
    """ارسال پیام فعال‌سازی سرویس"""
    pkg_id = payment['package_id']
    pkg = PACKAGES.get(pkg_id)
    
    if not pkg:
        return
    
    # ساخت سرویس
    marzban_username = generate_username(user_id, None, None)
    result = await marzban.create_user(marzban_username, pkg['traffic'], pkg['duration'])
    
    if result:
        from datetime import datetime, timedelta
        expire_date = datetime.now() + timedelta(days=pkg['duration'])
        create_order(user_id, pkg_id, marzban_username, pkg['price'], expire_date, result['subscription_url'])
        
        text = f"✅ <b>پرداخت موفق!</b>\n\n"
        text += f"📦 پکیج: {pkg['name']}\n"
        text += f"💰 مبلغ: {format_price(pkg['price'])}\n"
        text += f"🔢 کد پیگیری: <code>{ref_id}</code>\n\n"
        text += f"👤 نام کاربری: <code>{marzban_username}</code>\n"
        text += f"📊 حجم: {format_bytes(pkg['traffic'])}\n"
        text += f"📅 تاریخ انقضا: {format_date(expire_date)}\n\n"
        text += f"🔗 لینک اتصال:\n<code>{result['subscription_url']}</code>"
        
        keyboard = [[InlineKeyboardButton("🏠 بازگشت به منو", callback_data="back_to_main")]]
        
        await application.bot.send_message(
            chat_id=user_id,
            text=text,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )


async def send_wallet_charge_message(user_id, payment, ref_id):
    """ارسال پیام شارژ کیف پول"""
    update_user_balance(user_id, payment['amount'], f"شارژ آنلاین - کد: {ref_id}")
    
    text = f"✅ <b>شارژ موفق!</b>\n\n"
    text += f"💰 مبلغ: {format_price(payment['amount'])}\n"
    text += f"🔢 کد پیگیری: <code>{ref_id}</code>\n\n"
    text += f"💵 موجودی جدید: {format_price(get_user(user_id)['balance'])}"
    
    await application.bot.send_message(
        chat_id=user_id,
        text=text,
        parse_mode='HTML'
    )


async def start_webserver():
    """راه‌اندازی وب‌سرور برای callback"""
    app = web.Application()
    app.router.add_get('/zarinpal/callback', zarinpal_callback)
    
    runner = web.AppRunner(app)
    await runner.setup()
    
    site = web.TCPSite(runner, '0.0.0.0', 8081)
    await site.start()
    
    logger.info("🌐 وب‌سرور callback روی پورت 8081 راه‌اندازی شد")


# ==================== MAIN ====================

async def main():
    """✅ Run Bot + Web Server"""
    global application
    
    try:
        init_db()
        logger.info("✅ دیتابیس آماده شد")
    except Exception as e:
        logger.error(f"❌ خطا در init_db: {e}")
        return

    try:
        request = HTTPXRequest(
            connection_pool_size=20,
            connect_timeout=30.0,
            read_timeout=30.0,
            write_timeout=30.0,
            pool_timeout=30.0
        )

        application = Application.builder()\
            .token(TELEGRAM_TOKEN)\
            .request(request)\
            .build()

        application.add_handler(CommandHandler("start", start))
        application.add_handler(CallbackQueryHandler(button_handler))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))

        logger.info("✅ ربات راه‌اندازی شد")

        # شروع وب‌سرور
        asyncio.create_task(start_webserver())
        
        # شروع ربات
        await application.initialize()
        await application.start()
        await application.updater.start_polling()
        
        logger.info("✅ Bot + Web Server running...")
        
        # نگه‌داشتن
        await asyncio.Event().wait()

    except Exception as e:
        logger.error(f"❌ خطای اجرا: {e}")


if __name__ == '__main__':
    asyncio.run(main())


